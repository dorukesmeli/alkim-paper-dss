"""
app.py — Alkım Kağıt Karar Destek Sistemi (DSS)
Türkçe, modern Streamlit arayüzü — kağıt kesme optimizasyonu.
"""

import sys
import os
import io
import math
import time
import base64
import traceback

import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# ---------------------------------------------------------------------------
# Page config  — MUST be first Streamlit call
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Alkım Kağıt DSS",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Custom CSS — Premium dark-sidebar / light-main design
# ---------------------------------------------------------------------------
CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ── Global reset ─────────────────────────────────────────────── */
html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
#MainMenu                             { visibility: hidden !important; }
footer                                { visibility: hidden !important; }
header                                { visibility: hidden !important; height: 0 !important; }
[data-testid="stDecoration"]          { display: none !important; visibility: hidden !important; }
[data-testid="stStatusWidget"]        { display: none !important; visibility: hidden !important; }
[data-testid="stToolbar"]             { display: none !important; visibility: hidden !important; height: 0 !important; }
[data-testid="stDeployButton"]        { display: none !important; visibility: hidden !important; }
[data-testid="stBaseButton-headerNoPadding"] { display: none !important; visibility: hidden !important; }
.stDeployButton                       { display: none !important; }
/* Streamlit Cloud floating badge / viewer / manage-app fallback selectors */
div[class*="viewerBadge"]             { display: none !important; visibility: hidden !important; }
div[class*="stAppDeployButton"]       { display: none !important; visibility: hidden !important; }
div[class*="toolbar"]                 { display: none !important; visibility: hidden !important; }
div[class*="decoration"]              { display: none !important; visibility: hidden !important; }
div[class*="statusWidget"]            { display: none !important; visibility: hidden !important; }
/* ── Sidebar always open — fixed position, internal scroll only ─ */
[data-testid="stSidebarCollapseButton"]{ display: none !important; }
[data-testid="collapsedControl"]       { display: none !important; }
section[data-testid="stSidebar"] {
    position: sticky !important;
    top: 0 !important;
    height: 100vh !important;
    align-self: flex-start !important;
    overflow-y: auto !important;
    overflow-x: hidden !important;
    width: 260px !important;
    min-width: 260px !important;
    max-width: 260px !important;
    transform: translateX(0px) !important;
    visibility: visible !important;
    display: flex !important;
    flex-direction: column !important;
    flex-shrink: 0 !important;
}

/* ── App background ───────────────────────────────────────────── */
.stApp { background: #F4F5F7; }
.block-container {
    padding: 1rem 1.8rem 2rem 1.8rem !important;
    max-width: 1400px !important;
}

/* ── Sidebar ──────────────────────────────────────────────────── */
[data-testid="stSidebar"] {
    background: #111827 !important;
    border-right: 1px solid #1F2937;
}
[data-testid="stSidebarContent"] {
    padding: 1.5rem 1rem 2.5rem 1rem;
    overflow-y: auto;
    overflow-x: hidden;
    height: 100%;
    display: flex;
    flex-direction: column;
}

/* Sidebar text */
[data-testid="stSidebarContent"] p,
[data-testid="stSidebarContent"] label,
[data-testid="stSidebarContent"] span,
[data-testid="stSidebarContent"] div { color: #D1D5DB !important; }

/* Sidebar nav buttons */
[data-testid="stSidebarContent"] .stButton > button {
    width: 100% !important;
    background: transparent !important;
    color: #9CA3AF !important;
    border: none !important;
    border-radius: 8px !important;
    text-align: left !important;
    padding: 10px 14px !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    transition: all 0.18s ease !important;
    margin-bottom: 2px !important;
}
[data-testid="stSidebarContent"] .stButton > button:hover {
    background: #1F2937 !important;
    color: #F9FAFB !important;
}
[data-testid="stSidebarContent"] .stButton > button[kind="primary"] {
    background: rgba(200, 16, 46, 0.18) !important;
    color: #F87171 !important;
    border-left: 3px solid #C8102E !important;
}

/* ── Main headings ────────────────────────────────────────────── */
h1 { font-size: 1.75rem !important; font-weight: 700 !important; color: #111827 !important; }
h2 { font-size: 1.3rem  !important; font-weight: 600 !important; color: #1F2937 !important; }
h3 { font-size: 1.05rem !important; font-weight: 600 !important; color: #374151 !important; }

/* ── Cards ────────────────────────────────────────────────────── */
.dss-card {
    background: #FFFFFF;
    border-radius: 14px;
    padding: 18px 22px;
    box-shadow: 0 1px 3px rgba(0,0,0,.08), 0 4px 12px rgba(0,0,0,.05);
    margin-bottom: 0.8rem;
}
.dss-card-red {
    border-left: 4px solid #C8102E;
}
.dss-card-gray {
    border-left: 4px solid #6B7280;
}

/* ── Metric card ──────────────────────────────────────────────── */
.metric-card {
    background: #FFFFFF;
    border-radius: 12px;
    padding: 20px 22px;
    box-shadow: 0 1px 3px rgba(0,0,0,.07), 0 3px 10px rgba(0,0,0,.05);
    text-align: center;
    border-top: 3px solid #C8102E;
    height: 100%;
    min-height: 110px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    box-sizing: border-box;
}
.metric-card .metric-value {
    font-size: 2rem;
    font-weight: 700;
    color: #111827;
    line-height: 1.1;
}
.metric-card .metric-label {
    font-size: 0.78rem;
    font-weight: 500;
    color: #6B7280;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin-top: 6px;
}
.metric-card .metric-sub {
    font-size: 0.8rem;
    color: #9CA3AF;
    margin-top: 2px;
}

/* ── Equal-height bordered panel pairs ───────────────────────── */
/* Makes every st.columns row that contains st.container(border=True)
   stretch both columns to the same height as the tallest one.        */
[data-testid="stHorizontalBlock"]:has(
  > [data-testid="stColumn"] [data-testid="stVerticalBlockBorderWrapper"]
) {
    align-items: stretch !important;
}
[data-testid="stHorizontalBlock"]:has(
  > [data-testid="stColumn"] [data-testid="stVerticalBlockBorderWrapper"]
) > [data-testid="stColumn"] {
    display: flex !important;
    flex-direction: column !important;
}
[data-testid="stHorizontalBlock"]:has(
  > [data-testid="stColumn"] [data-testid="stVerticalBlockBorderWrapper"]
) > [data-testid="stColumn"] > [data-testid="stVerticalBlock"] {
    flex: 1 !important;
    display: flex !important;
    flex-direction: column !important;
}
[data-testid="stHorizontalBlock"]:has(
  > [data-testid="stColumn"] [data-testid="stVerticalBlockBorderWrapper"]
) [data-testid="stVerticalBlockBorderWrapper"] {
    flex: 1 !important;
    min-height: 182px;
}

/* ── Hero ─────────────────────────────────────────────────────── */
.hero-banner {
    background: linear-gradient(135deg, #111827 0%, #1F2937 60%, #7f1d1d 100%);
    border-radius: 16px;
    padding: 40px 48px;
    color: #FFFFFF;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.hero-banner::after {
    content: "";
    position: absolute;
    top: -40px; right: -40px;
    width: 220px; height: 220px;
    background: rgba(200,16,46,0.15);
    border-radius: 50%;
}
.hero-title {
    font-size: 2.1rem;
    font-weight: 800;
    letter-spacing: -0.02em;
    line-height: 1.2;
    margin-bottom: 10px;
}
.hero-subtitle {
    font-size: 1rem;
    color: #D1D5DB;
    line-height: 1.55;
    max-width: 600px;
}
.hero-badge {
    display: inline-block;
    background: rgba(200,16,46,0.25);
    color: #FCA5A5;
    border: 1px solid rgba(200,16,46,0.4);
    border-radius: 20px;
    padding: 4px 14px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.05em;
    margin-bottom: 14px;
}

/* ── Method cards ─────────────────────────────────────────────── */
.method-card {
    background: #FFFFFF;
    border: 2px solid #E5E7EB;
    border-radius: 12px;
    padding: 18px 20px;
    cursor: pointer;
    transition: all 0.2s;
    margin-bottom: 10px;
}
.method-card:hover { border-color: #C8102E; box-shadow: 0 0 0 3px rgba(200,16,46,0.08); }
.method-card.selected { border-color: #C8102E; background: #FFF5F5; }
.method-card .method-title { font-weight: 600; color: #111827; font-size: 0.95rem; }
.method-card .method-desc  { font-size: 0.8rem; color: #6B7280; margin-top: 4px; }
.recommended-badge {
    display: inline-block;
    background: #C8102E;
    color: white;
    border-radius: 6px;
    padding: 2px 8px;
    font-size: 0.7rem;
    font-weight: 600;
    margin-left: 8px;
    vertical-align: middle;
}

/* ── Section header ───────────────────────────────────────────── */
.section-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 0.5rem;
    padding-bottom: 10px;
    border-bottom: 2px solid #E5E7EB;
}
.section-header .icon {
    width: 34px; height: 34px;
    background: rgba(200,16,46,0.1);
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1rem;
}

/* ── Status badges ────────────────────────────────────────────── */
.badge-optimal  { background:#D1FAE5; color:#065F46; border-radius:6px; padding:3px 10px; font-size:.75rem; font-weight:600; }
.badge-timelimit{ background:#FEF3C7; color:#92400E; border-radius:6px; padding:3px 10px; font-size:.75rem; font-weight:600; }
.badge-error    { background:#FEE2E2; color:#991B1B; border-radius:6px; padding:3px 10px; font-size:.75rem; font-weight:600; }
.badge-feasible { background:#DBEAFE; color:#1E40AF; border-radius:6px; padding:3px 10px; font-size:.75rem; font-weight:600; }

/* ── Warning / info banners ───────────────────────────────────── */
.warn-banner {
    background: #FFFBEB;
    border: 1px solid #FCD34D;
    border-left: 4px solid #F59E0B;
    border-radius: 10px;
    padding: 14px 18px;
    margin-bottom: 1rem;
    color: #78350F;
    font-size: 0.88rem;
}
.info-banner {
    background: #EFF6FF;
    border: 1px solid #BFDBFE;
    border-left: 4px solid #3B82F6;
    border-radius: 10px;
    padding: 14px 18px;
    margin-bottom: 1rem;
    color: #1E3A5F;
    font-size: 0.88rem;
}
.err-banner {
    background: #FEF2F2;
    border: 1px solid #FECACA;
    border-left: 4px solid #EF4444;
    border-radius: 10px;
    padding: 14px 18px;
    margin-bottom: 1rem;
    color: #7F1D1D;
    font-size: 0.88rem;
}

/* ── Divider ──────────────────────────────────────────────────── */
.dss-divider { border: none; border-top: 1px solid #E5E7EB; margin: 0.9rem 0; }

/* ── Primary button ───────────────────────────────────────────── */
.stButton > button[kind="primary"] {
    background: #C8102E !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    padding: 10px 24px !important;
    font-size: 0.95rem !important;
    transition: background 0.2s !important;
}
.stButton > button[kind="primary"]:hover {
    background: #A50E26 !important;
}

/* ── Dataframe ────────────────────────────────────────────────── */
[data-testid="stDataFrame"] > div { border-radius: 10px !important; overflow: hidden; }

/* ── Expander ─────────────────────────────────────────────────── */
[data-testid="stExpander"] {
    border: 1px solid #E5E7EB !important;
    border-radius: 10px !important;
    background: #FFFFFF !important;
}

/* ── Slider ───────────────────────────────────────────────────── */
[data-testid="stSlider"] [role="slider"] { background: #C8102E !important; }
[data-testid="stSlider"] [data-testid="stSliderTrack"] > div:first-child { background: #C8102E !important; }

/* ── Tabs ─────────────────────────────────────────────────────── */
[data-testid="stTabs"] [data-baseweb="tab-list"] {
    background: #FFFFFF;
    border-radius: 10px;
    padding: 4px;
    gap: 2px;
}
[data-testid="stTabs"] [data-baseweb="tab"] {
    border-radius: 8px !important;
    font-weight: 500 !important;
    color: #6B7280 !important;
}
[data-testid="stTabs"] [aria-selected="true"] {
    background: #C8102E !important;
    color: #FFFFFF !important;
}

/* ── Number input ─────────────────────────────────────────────── */
[data-testid="stNumberInput"] input { border-radius: 8px !important; }

/* ── Select box ───────────────────────────────────────────────── */
[data-testid="stSelectbox"] > div > div {
    border-radius: 8px !important;
}
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Local imports (after page_config)
# ---------------------------------------------------------------------------
sys.path.insert(0, os.path.dirname(__file__))
from data_utils import (
    validate_orders, validate_params,
    estimate_reel_count, estimate_reels_detailed, build_data_dict,
    build_roll_display_df, get_sample_orders_df,
    normalize_orders_df,
)
from solver_wrappers import run_selected_solver


# ---------------------------------------------------------------------------
# UI number formatting helpers
# ---------------------------------------------------------------------------
def fmt_n(v) -> str:
    """Format a number for display: whole numbers show no decimal, others strip trailing zeros."""
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else f"{f:g}"
    except Exception:
        return str(v)

def fmt_cm(v) -> str:
    """Format a length value with ' cm' suffix, no unnecessary decimals."""
    return f"{fmt_n(v)} cm"


# ---------------------------------------------------------------------------
# Session state init
# ---------------------------------------------------------------------------
DEFAULTS = {
    "active_page": "home",
    "orders_df": None,
    "method": "Heuristic Model 2",
    "lambda_val": 0.5,
    "L": 348.0,
    "B": 6,
    "CT": 5.0,
    "time_limit": 3600,
    "tj_table": None,           # for Model 1 / Heuristic 1
    "solver_result": None,
    "type_map": None,
    "safe_count": None,
    "data_dict": None,
    "optimization_step": 0,
    "priority_mode": None,          # None | "paper_waste" | "time_waste" | "balanced"
    "strategy_mode": None,          # None | "given_schedule" | "auto_type" | "full_auto"
    "given_schedule_df": None,      # DataFrame: bobin_no, paper_type
    "auto_type_counts": None,       # dict: {paper_type_str: bobin_count} for auto_type strategy
}

for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ---------------------------------------------------------------------------
# Paper-type mapping  (frontend ↔ backend)
# ---------------------------------------------------------------------------

PAPER_TYPE_LABELS: dict[int, str] = {
    1: "1.HAMUR OFSET",
    2: "1.HAMUR IVORY(D)",
    3: "1.HAMUR BARDAK",
    4: "1.HAMUR AMBALAJ KAGIDI",
    5: "FSC MIX CREDIT KRAFT ÇANTA OB",
    6: "1.HAMUR BRISTOL",
    7: "1.HAMUR KRAFT CANTA",
    8: "1.HAMUR CEVAP KAGIDI",
}

PAPER_TYPE_IDS: dict[str, int] = {v: k for k, v in PAPER_TYPE_LABELS.items()}

# Ordered list used for dropdowns
PAPER_TYPE_OPTIONS: list[str] = [PAPER_TYPE_LABELS[k] for k in sorted(PAPER_TYPE_LABELS)]

# Short codes shown inside narrow diagram segments
PAPER_TYPE_SHORT: dict[str, str] = {
    "1.HAMUR OFSET":                "OFSET",
    "1.HAMUR IVORY(D)":             "IVORY",
    "1.HAMUR BARDAK":               "BARDAK",
    "1.HAMUR AMBALAJ KAGIDI":       "AMBALAJ",
    "FSC MIX CREDIT KRAFT ÇANTA OB":"KRT.ÇNTA",
    "1.HAMUR BRISTOL":              "BRISTOL",
    "1.HAMUR KRAFT CANTA":          "KRF.ÇNTA",
    "1.HAMUR CEVAP KAGIDI":         "CEVAP",
}


def pt_label(val) -> str:
    """Return user-facing label for any paper_type value (int, numeric str, or label)."""
    if isinstance(val, str):
        s = val.strip()
        if s in PAPER_TYPE_IDS:
            return s                          # already a valid label
        try:
            return PAPER_TYPE_LABELS.get(int(s), s)   # "1" → "1.HAMUR OFSET"
        except (ValueError, TypeError):
            return s
    try:
        return PAPER_TYPE_LABELS.get(int(val), str(val))
    except (ValueError, TypeError):
        return str(val)


def pt_normalize_series(series: "pd.Series") -> "pd.Series":
    """Apply pt_label to every element of a paper_type Series."""
    return series.apply(pt_label)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def logo_b64() -> str:
    logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return ""


def metric_card(value, label, sub=None, color="#C8102E"):
    sub_html = f'<div class="metric-sub">{sub}</div>' if sub else ""
    st.markdown(
        f'<div class="metric-card" style="border-top-color:{color}">'
        f'<div class="metric-value">{value}</div>'
        f'<div class="metric-label">{label}</div>'
        f'{sub_html}'
        f'</div>',
        unsafe_allow_html=True,
    )


def section_header(icon, title, subtitle=None):
    sub_html = f"<p style='color:#6B7280;font-size:.85rem;margin:4px 0 0 0'>{subtitle}</p>" if subtitle else ""
    st.markdown(
        f"""<div class="section-header">
                <div class="icon">{icon}</div>
                <div><h2 style='margin:0'>{title}</h2>{sub_html}</div>
            </div>""",
        unsafe_allow_html=True,
    )


def nav_to(page: str):
    st.session_state.active_page = page


def status_badge(status: str) -> str:
    badges = {
        "optimal":               '<span class="badge-optimal">✓ Optimal</span>',
        "time_limit":            '<span class="badge-timelimit">⏱ Zaman Limiti</span>',
        "feasible":              '<span class="badge-feasible">✓ Uygun Çözüm</span>',
        "infeasible":            '<span class="badge-error">✗ Çözüm Yok</span>',
        "time_limit_no_solution":'<span class="badge-error">✗ Zaman Limiti (Çözüm Yok)</span>',
        "error":                 '<span class="badge-error">✗ Hata</span>',
    }
    return badges.get(status, f'<span class="badge-feasible">{status}</span>')


# ---------------------------------------------------------------------------
# Sidebar navigation
# ---------------------------------------------------------------------------
NAV_PAGES = [
    ("home",         "🏠", "Ana Sayfa"),
    ("optimization", "✂️",  "Kesim Optimizasyonu"),
    ("results",      "📊", "Sonuçlar"),
    ("analysis",     "📈", "Görsel Analiz"),
    ("export",       "💾", "Dışa Aktarma"),
]


def render_sidebar():
    with st.sidebar:
        logo = logo_b64()
        if logo:
            st.markdown(
                f'<img src="data:image/png;base64,{logo}" style="width:160px;margin-bottom:16px;display:block;">',
                unsafe_allow_html=True,
            )
        st.markdown(
            '<p style="color:#6B7280;font-size:0.72rem;letter-spacing:.08em;text-transform:uppercase;margin-bottom:1rem">Karar Destek Sistemi</p>',
            unsafe_allow_html=True,
        )

        for page_id, icon, label in NAV_PAGES:
            cur = st.session_state.active_page
            is_active = (cur == page_id) or (page_id == "optimization" and cur in ("orders", "settings"))
            btn_type = "primary" if is_active else "secondary"
            if st.button(f"{icon}  {label}", key=f"nav_{page_id}",
                         use_container_width=True, type=btn_type):
                if page_id == "optimization":
                    st.session_state.optimization_step = 0
                nav_to(page_id)
                st.rerun()

        st.markdown("<hr style='border-color:#1F2937;margin:1.5rem 0'>", unsafe_allow_html=True)

        # Quick status
        has_orders = st.session_state.orders_df is not None
        has_result = st.session_state.solver_result is not None

        def dot(ok): return "🟢" if ok else "🔴"
        st.markdown(
            f"""<div style='font-size:.78rem;color:#6B7280;line-height:1.9'>
                {dot(has_orders)} Siparişler yüklendi<br>
                {dot(has_result)} Çözüm mevcut
            </div>""",
            unsafe_allow_html=True,
        )

        st.markdown("<hr style='border-color:#1F2937;margin:1.5rem 0'>", unsafe_allow_html=True)
        st.markdown(
            '<p style="color:#374151;font-size:.7rem;text-align:center">Alkım Kağıt A.Ş. © 2026<br>Bitirme Projesi</p>',
            unsafe_allow_html=True,
        )


# ===========================================================================
# PAGE: ANA SAYFA
# ===========================================================================
def page_home():
    logo = logo_b64()
    logo_html = (
        f'<img src="data:image/png;base64,{logo}" style="height:52px;margin-bottom:18px;display:block;">'
        if logo else ""
    )
    st.markdown(
        f"""<div class="hero-banner">
            {logo_html}
            <div class="hero-badge">KARAR DESTEK SİSTEMİ</div>
            <div class="hero-title">Bobin Kesme Optimizasyonu</div>
            <div class="hero-subtitle">
                Alkım Kağıt'ın jumbo bobin kesme süreçlerini akıllı algoritmalar ile optimize edin.
                Kağıt kaybını ve zaman kaybını minimum düzeye indirerek üretim verimliliğini artırın.
            </div>
        </div>""",
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(
            """<div class="dss-card dss-card-red">
                <div style="font-size:1.8rem;margin-bottom:10px">✂️</div>
                <h3 style="margin:0 0 6px 0">Kesim Optimizasyonu</h3>
                <p style="color:#6B7280;font-size:.87rem;margin:0">
                    Jumbo bobinlerin en verimli biçimde kesilmesi için matematiksel optimizasyon ve
                    sezgisel algoritmalar kullanır.
                </p>
            </div>""",
            unsafe_allow_html=True,
        )
    with col2:
        st.markdown(
            """<div class="dss-card dss-card-red">
                <div style="font-size:1.8rem;margin-bottom:10px">📉</div>
                <h3 style="margin:0 0 6px 0">İki Boyutlu Hedef</h3>
                <p style="color:#6B7280;font-size:.87rem;margin:0">
                    Kağıt kaybı ve zaman kaybı birlikte minimize edilir.
                    Lambda parametresi ile öncelik belirlenir.
                </p>
            </div>""",
            unsafe_allow_html=True,
        )

    st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)

    _hw_l, _hw_r = st.columns([3, 2])
    with _hw_l:
        st.markdown(
            """<div class="dss-card">
                <h3>🚀 Nasıl Başlanır?</h3>
                <ol style="color:#374151;font-size:.88rem;line-height:1.9">
                    <li><b>Kesim Optimizasyonuna Başla</b> — Ana sayfadaki butonla kesim planı oluşturma sürecine girin.</li>
                    <li><b>Sipariş Girişi</b> — Siparişlerinizi manuel olarak girin veya Excel/CSV dosyası yükleyin.</li>
                    <li><b>Öncelik Tercihi</b> — Çözümün kağıt kaybına mı, zaman kaybına mı yoksa dengeli bir yapıya mı odaklanacağını seçin.</li>
                    <li><b>Kesim Planı Stratejisi</b> — Üretim sırasını kendiniz verebilir, mevcut tip dağılımına göre planlama yapabilir veya tüm planlamayı sisteme bırakabilirsiniz.</li>
                    <li><b>Optimizasyonu Başlat</b> — Sistem seçilen ayarlara göre kesim planını oluşturur.</li>
                    <li><b>Sonuçları İncele</b> — Bobin bazlı karar tablosunu, görsel analizleri ve dışa aktarma seçeneklerini inceleyin.</li>
                </ol>
            </div>""",
            unsafe_allow_html=True,
        )
    with _hw_r:
        st.markdown(
            """<div class="dss-card dss-card-red">
                <div style="font-size:1.8rem;margin-bottom:10px">🎯</div>
                <h3 style="margin:0 0 6px 0">3 Planlama Stratejisi</h3>
                <ul style="color:#6B7280;font-size:.87rem;margin:0;padding-left:1.1rem;line-height:1.9">
                    <li><b>Verilen Sıraya Göre</b> — Üretim sırasını kendiniz belirleyin.</li>
                    <li><b>Mevcut Tip Dağılımına Göre</b> — Tip bazlı bobin adetini siz verin, sırayı sistem kursun.</li>
                    <li><b>Tam Otomatik</b> — Her şeyi sisteme bırakın.</li>
                </ul>
            </div>""",
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)
    col_btn, _ = st.columns([1, 3])
    with col_btn:
        if st.button("✂️  Kesim Optimizasyonuna Başla →", type="primary", use_container_width=True):
            st.session_state.optimization_step = 0
            nav_to("optimization")
            st.rerun()



# ===========================================================================
# PAGE: KESİM OPTİMİZASYONU — 3-step workflow
# ===========================================================================

def _opt_progress_bar(active_step: int):
    """Horizontal step progress indicator — active step in Alkım red."""
    steps = [
        ("1", "Sipariş Girişi"),
        ("2", "Öncelik Tercihi"),
        ("3", "Kesim Planı Stratejisi"),
    ]
    parts = []
    for i, (num, label) in enumerate(steps):
        if i + 1 == active_step:
            style = "color:#C8102E;font-weight:700;font-size:.95rem"
            dot = (
                'style="display:inline-block;width:10px;height:10px;border-radius:50%;'
                'background:#C8102E;margin-right:5px;vertical-align:middle"'
            )
            parts.append(f'<span {dot}></span><span style="{style}">{num}. {label}</span>')
        elif i + 1 < active_step:
            style = "color:#374151;font-weight:500;font-size:.9rem"
            parts.append(f'<span style="{style}">✓ {num}. {label}</span>')
        else:
            style = "color:#9CA3AF;font-weight:400;font-size:.9rem"
            parts.append(f'<span style="{style}">{num}. {label}</span>')
        if i < len(steps) - 1:
            parts.append('<span style="color:#D1D5DB;padding:0 10px">→</span>')
    st.markdown(
        '<div style="background:#F9FAFB;border:1px solid #E5E7EB;border-radius:10px;' +
        'padding:12px 20px;margin-bottom:1.5rem;display:flex;align-items:center;gap:4px">' +
        "".join(parts) + "</div>",
        unsafe_allow_html=True,
    )


def _step_nav(back_label=None, back_step=None, next_label=None, next_step=None,
               next_disabled=False, next_key="opt_next", back_key="opt_back"):
    """Renders back/forward navigation buttons for a workflow step."""
    st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)
    cols = st.columns([1, 3, 1]) if back_label and next_label else (
        st.columns([4, 1]) if next_label else st.columns([1, 4])
    )
    if back_label and next_label:
        col_b, _, col_n = cols
    elif next_label:
        _, col_n = cols
        col_b = None
    else:
        col_b, _ = cols
        col_n = None

    if col_b and back_label:
        with col_b:
            if st.button(f"← {back_label}", use_container_width=True, key=back_key):
                st.session_state.optimization_step = back_step
                st.rerun()
    if col_n and next_label:
        with col_n:
            if st.button(next_label + " →", type="primary",
                         use_container_width=True, key=next_key,
                         disabled=next_disabled):
                st.session_state.optimization_step = next_step
                st.rerun()


def page_optimization():
    step = st.session_state.get("optimization_step", 0)

    # ══════════════════════════════════════════════════════════════════════════
    if step == 0:
        # Landing page — 3 cards + start button
        # ══════════════════════════════════════════════════════════════════════
        section_header("✂️", "Kesim Optimizasyonu",
                       "Jumbo bobin kesme planını oluşturmak için adımları takip edin")
        st.markdown("<br>", unsafe_allow_html=True)

        # All three cards in one markdown block so flexbox can equalise their heights.
        # align-items:stretch makes every card as tall as the tallest sibling.
        # Each card uses the same CARD_BASE; only background differs per card.
        CARD_BASE = (
            "flex:1 1 0;border-radius:12px;padding:1.5rem;"
            "display:flex;flex-direction:column;gap:12px;"
            "box-sizing:border-box;margin:0;text-align:left;"
        )
        ICON_STYLE  = "font-size:1.8rem;line-height:1;margin:0;padding:0;text-align:left;"
        TITLE_STYLE = "font-size:1.05rem;font-weight:700;line-height:1.3;margin:0;padding:0;text-align:left;"
        DESC_STYLE  = "font-size:.88rem;line-height:1.5;margin:0;padding:0;flex-grow:1;text-align:left;"

        # Card colour definitions: (background, title-colour, desc-colour)
        C1 = ("#111827", "white", "#D1D5DB")   # Dark Navy
        C2 = ("#C8102E", "white", "#FEE2E2")   # Alkım Red
        C3 = ("#1F2937", "white", "#D1D5DB")   # Dark Gray

        def _card(bg, tc, dc, icon, title, desc):
            return (
                f'<div style="{CARD_BASE}background:{bg};">'
                f'<div style="{ICON_STYLE}">{icon}</div>'
                f'<div style="{TITLE_STYLE}color:{tc};">{title}</div>'
                f'<div style="{DESC_STYLE}color:{dc};">{desc}</div>'
                f'</div>'
            )

        st.markdown(
            '<div style="display:flex;gap:16px;align-items:stretch;width:100%;margin:0;padding:0">'
            + _card(*C1, "📋", "1. Adım: Sipariş Girişi",
                    "Müşteri siparişlerini manuel olarak girin veya Excel/CSV dosyası yükleyerek sisteme aktarın.")
            + _card(*C2, "⚖️", "2. Adım: Öncelik Tercihi",
                    "Çözümün kağıt kaybına mı yoksa zaman kaybına mı daha fazla odaklanacağını belirleyin.")
            + _card(*C3, "🎯", "3. Adım: Kesim Planı Stratejisi",
                    "Bobin tiplerinin önceden verildiği veya sistem tarafından otomatik belirlendiği çözüm yaklaşımını seçin.")
            + "</div>",
            unsafe_allow_html=True,
        )

        st.markdown("<br>", unsafe_allow_html=True)
        col_btn, _ = st.columns([1, 2])
        with col_btn:
            if st.button("✂️  Kesim Optimizasyonuna Başla", type="primary",
                         use_container_width=True, key="opt_landing_start"):
                st.session_state.optimization_step = 1
                st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    elif step == 1:
        # Step 1 — Sipariş Girişi
        # ══════════════════════════════════════════════════════════════════════
        _opt_progress_bar(1)
        section_header("📋", "1. Adım: Sipariş Girişi",
                       "Müşteri siparişlerini girin veya dosya yükleyin")

        tab_manual, tab_upload = st.tabs(["✏️  Manuel Giriş", "📁  Excel / CSV Yükleme"])

        with tab_manual:
            st.markdown("##### Sipariş Tablosu")
            st.caption("Satır eklemek için son satırın altına tıklayın. Kağıt Tipi sütunundan gerçek kağıt cinsini seçin.")

            if st.session_state.orders_df is not None:
                init_df = normalize_orders_df(st.session_state.orders_df)
                init_df["paper_type"] = pt_normalize_series(init_df["paper_type"])
                for col in ["paper_type", "length", "quantity"]:
                    if col not in init_df.columns:
                        init_df[col] = ""
            else:
                init_df = pd.DataFrame({
                    "paper_type": pd.Series([], dtype=str),
                    "length":     pd.Series([], dtype=float),
                    "quantity":   pd.array([], dtype="Int64"),
                })

            edited = st.data_editor(
                init_df[["paper_type", "length", "quantity"]],
                num_rows="dynamic",
                use_container_width=True,
                column_config={
                    "paper_type": st.column_config.SelectboxColumn(
                        "Kağıt Tipi", options=PAPER_TYPE_OPTIONS, required=True,
                    ),
                    "length":     st.column_config.NumberColumn("Uzunluk (cm)", min_value=0.1, format="%.1f"),
                    "quantity":   st.column_config.NumberColumn("Miktar (adet)", min_value=1, step=1, format="%d"),
                },
                hide_index=True,
                key="opt_orders_editor",
            )

            col_a, col_b, col_c = st.columns([1, 1, 2])
            with col_a:
                if st.button("✅  Siparişleri Kaydet", type="primary", use_container_width=True,
                             key="opt_save_orders"):
                    df_ed = normalize_orders_df(edited.dropna(how="all").reset_index(drop=True))
                    df_ed["paper_type"] = pt_normalize_series(df_ed["paper_type"])
                    ok, errs = validate_orders(df_ed)
                    if ok:
                        st.session_state.orders_df = df_ed
                        st.session_state.solver_result = None
                        st.success(f"✓ {len(df_ed)} sipariş kaydedildi.")
                    else:
                        for e in errs:
                            st.error(e)
            with col_b:
                if st.button("🔄  Örnek Veriyi Yükle", use_container_width=True,
                             key="opt_sample_data"):
                    _csv_path = os.path.join(
                        os.path.dirname(os.path.abspath(__file__)),
                        "example_orders.csv",
                    )
                    _s = get_sample_orders_df(_csv_path)
                    _s["paper_type"] = pt_normalize_series(_s["paper_type"])
                    st.session_state.orders_df = _s
                    st.session_state.solver_result = None
                    st.rerun()
            with col_c:
                try:
                    import io as _tio
                    from openpyxl import Workbook as _WB2
                    from openpyxl.styles import Font as _Fnt2, PatternFill as _PF2, Alignment as _Al2
                    from openpyxl.worksheet.datavalidation import DataValidation as _DV2
                    _tbuf2 = _tio.BytesIO()
                    _twb2  = _WB2()
                    _tws2  = _twb2.active
                    _tws2.title = "Siparişler"
                    _rf2 = _PF2(start_color="C8102E", end_color="C8102E", fill_type="solid")
                    _hf2 = _Fnt2(bold=True, color="FFFFFF", size=11)
                    from openpyxl.styles import Alignment as _AlI2
                    for _ci2, _h2 in enumerate(["Kağıt Tipi", "Uzunluk", "Miktar"], start=1):
                        _c2 = _tws2.cell(row=1, column=_ci2, value=_h2)
                        _c2.fill = _rf2; _c2.font = _hf2
                        _c2.alignment = _AlI2(horizontal="center")
                    _ex2 = [
                        ("1.HAMUR OFSET", 140.0, 5),
                        ("1.HAMUR IVORY(D)", 170.0, 4),
                    ]
                    for _ri2, (_pt2, _ln2, _qty2) in enumerate(_ex2, start=2):
                        _tws2.cell(row=_ri2, column=1, value=_pt2)
                        _tws2.cell(row=_ri2, column=2, value=_ln2)
                        _tws2.cell(row=_ri2, column=3, value=_qty2)
                    # Bilgi notu — tablonun sağına (E1:G3), açık sarı arka plan
                    from openpyxl.styles import PatternFill as _NFill2, Font as _NFont2, Alignment as _NAl2
                    _note_txt2 = (
                        "⚠  Bu satırlar yalnızca örnek veri formatını göstermek için eklenmiştir.\n"
                        "Kendi siparişlerinizi girmeden önce bu örnek satırları silin\n"
                        "ve yerine gerçek verilerinizi yazın."
                    )
                    _tws2.merge_cells("E1:G3")
                    _nc2 = _tws2["E1"]
                    _nc2.value = _note_txt2
                    _nc2.fill = _NFill2(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                    _nc2.font = _NFont2(size=9, italic=True, color="5C4A00")
                    _nc2.alignment = _NAl2(wrap_text=True, vertical="top", horizontal="left")
                    # Dropdown validation for Kağıt Tipi column
                    _dv_formula2 = '"' + ",".join(PAPER_TYPE_OPTIONS) + '"'
                    _dv2 = _DV2(type="list", formula1=_dv_formula2, sqref="A2:A1000",
                                showDropDown=False, showErrorMessage=True,
                                error="Lütfen listeden bir kağıt tipi seçin.",
                                errorTitle="Geçersiz Kağıt Tipi")
                    _tws2.add_data_validation(_dv2)
                    from openpyxl.utils import get_column_letter as _gcl2
                    for _ci3, _w3 in [(1, 36), (2, 14), (3, 14), (4, 4), (5, 18), (6, 18), (7, 18)]:
                        _tws2.column_dimensions[_gcl2(_ci3)].width = _w3
                    _twb2.save(_tbuf2); _tbuf2.seek(0)
                    st.download_button(
                        "⬇️  Boş Şablon İndir (.xlsx)", data=_tbuf2,
                        file_name="siparis_sablonu.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception:
                    _tcsv2 = (
                        "Kağıt Tipi,Uzunluk,Miktar\n"
                        "1.HAMUR OFSET,140,5\n"
                        "1.HAMUR IVORY(D),170,4\n"
                    )
                    st.download_button(
                        "⬇️  Boş Şablon İndir (.csv)", data=_tcsv2,
                        file_name="siparis_sablonu.csv", mime="text/csv",
                        use_container_width=True,
                    )

        with tab_upload:
            st.markdown("##### Dosya Yükleme")
            st.caption("CSV veya Excel (xlsx) dosyası yükleyin. Gerekli sütunlar: **Kağıt Tipi**, **Uzunluk**, **Miktar**")
            uploaded = st.file_uploader("Dosya seçin", type=["csv", "xlsx", "xls"],
                                        label_visibility="collapsed", key="opt_file_upload")
            if uploaded:
                try:
                    if uploaded.name.endswith((".xlsx", ".xls")):
                        df_up = pd.read_excel(uploaded)
                    else:
                        df_up = pd.read_csv(uploaded)
                    df_up.columns = [c.strip().lower().replace(" ", "_") for c in df_up.columns]
                    _tr_to_en = {
                        "kağıt_tipi": "paper_type", "kagit_tipi": "paper_type",
                        "uzunluk": "length", "uzunluk_(cm)": "length",
                        "miktar": "quantity", "miktar_(adet)": "quantity",
                    }
                    df_up = df_up.rename(columns={k: v for k, v in _tr_to_en.items() if k in df_up.columns})
                    for _drop_col in ["order_id", "sipariş_numarası", "siparis_numarasi"]:
                        if _drop_col in df_up.columns:
                            df_up = df_up.drop(columns=[_drop_col])
                    df_up = normalize_orders_df(df_up.reset_index(drop=True))
                    # Normalize paper_type: numeric → label; validate known types
                    df_up["paper_type"] = pt_normalize_series(df_up["paper_type"])
                    _unk = [v for v in df_up["paper_type"].unique()
                            if str(v).strip() not in PAPER_TYPE_IDS and str(v).strip() != "nan"]
                    if _unk:
                        for _u in _unk:
                            st.error(f"Geçersiz kağıt tipi: '{_u}'. Lütfen şablondaki kağıt tiplerinden birini seçin.")
                    st.markdown("**Yüklenen veri önizlemesi:**")
                    st.dataframe(df_up.head(20), use_container_width=True, hide_index=True,
                        column_config={
                            "paper_type": st.column_config.TextColumn("Kağıt Tipi"),
                            "length":     st.column_config.NumberColumn("Uzunluk (cm)"),
                            "quantity":   st.column_config.NumberColumn("Miktar (adet)"),
                        })
                    ok, errs = validate_orders(df_up)
                    if ok and not _unk:
                        st.markdown(
                            f'<div class="info-banner">✓ {len(df_up)} sipariş doğrulandı ve hazır.</div>',
                            unsafe_allow_html=True,
                        )
                        if st.button("✅  Bu Veriyi Kullan", type="primary", key="opt_use_upload"):
                            st.session_state.orders_df = df_up
                            st.session_state.solver_result = None
                            st.success("Sipariş verisi kaydedildi.")
                    else:
                        for e in errs:
                            st.error(e)
                except Exception as e:
                    st.error(f"Dosya okunurken hata: {e}")

        # Current orders summary
        if st.session_state.orders_df is not None:
            df_cur = st.session_state.orders_df
            st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)
            st.markdown("#### 📌 Mevcut Siparişler")
            cc1, cc2 = st.columns(2)
            with cc1: metric_card(len(df_cur), "Toplam Sipariş Satırı")
            with cc2: metric_card(len(df_cur["paper_type"].unique()), "Farklı Kağıt Tipi")
            with st.expander("Sipariş tablosunu göster"):
                st.dataframe(df_cur, use_container_width=True, hide_index=True,
                    column_config={
                        "paper_type": st.column_config.TextColumn("Kağıt Tipi"),
                        "length":     st.column_config.NumberColumn("Uzunluk (cm)"),
                        "quantity":   st.column_config.NumberColumn("Miktar (adet)"),
                    })

        # Navigation
        has_orders = st.session_state.orders_df is not None
        if not has_orders:
            st.warning("⚠️ Devam etmek için önce sipariş girip kaydedin.")
        _step_nav(next_label="İleri: Öncelik Tercihi", next_step=2,
                  next_disabled=not has_orders, next_key="opt_1_next")

    # ══════════════════════════════════════════════════════════════════════════
    elif step == 2:
        # Step 2 — Öncelik Tercihi (lambda — kullanıcıya teknik gösterilmez)
        # ══════════════════════════════════════════════════════════════════════
        _opt_progress_bar(2)
        section_header("⚖️", "2. Adım: Öncelik Tercihi",
                       "Sistem hangi hedefe daha fazla önem versin?")

        st.markdown(
            '<div class="info-banner">'
            'Bu adımda sistemin kesim planını oluştururken hangi hedefe daha fazla önem vereceğini seçin. '
            '<b>Kağıt kaybını azaltmaya</b>, <b>zaman kaybını azaltmaya</b> veya '
            '<b>iki hedef arasında denge kurmaya</b> odaklanabilirsiniz.'
            '</div>',
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Mevcut seçimi oku (None = henüz seçim yapılmadı) ────────────────
        mode = st.session_state.get("priority_mode", None)

        # ── Kart stilleri ────────────────────────────────────────────────────
        def _card_wrap(selected: bool) -> str:
            if selected:
                return (
                    "border-radius:12px;padding:1.4rem 1.3rem 1rem 1.3rem;"
                    "box-sizing:border-box;"
                    "background:#FFF5F5;"
                    "border:2.5px solid #C8102E;"
                    "border-top:4px solid #C8102E;"
                )
            return (
                "border-radius:12px;padding:1.4rem 1.3rem 1rem 1.3rem;"
                "box-sizing:border-box;"
                "background:#FFFFFF;"
                "border:1.5px solid #E5E7EB;"
            )

        TITLE_S = "font-size:1.05rem;font-weight:700;color:#1E3A5F;margin:10px 0 8px 0"
        DESC_S  = "font-size:.87rem;color:#4B5563;line-height:1.65;margin:0"
        BADGE_S = (
            "display:inline-block;background:#C8102E;color:white;"
            "border-radius:6px;padding:4px 14px;font-size:.8rem;font-weight:600;margin-top:10px"
        )

        col1, col2, col3 = st.columns(3)

        # ── Kart 1: Kağıt Kaybını Azalt ─────────────────────────────────────
        with col1:
            is_pw = (mode == "paper_waste")
            st.markdown(
                f'<div style="{_card_wrap(is_pw)}">'
                '<div style="font-size:1.9rem">📉</div>'
                f'<div style="{TITLE_S}">Kağıt Kaybını Azalt</div>'
                f'<div style="{DESC_S}">Sistem, bobinlerde kullanılmadan kalan kağıt miktarını '
                'minimuma indirmeye odaklanır. Malzeme verimliliğini artırmak için uygundur.</div>'
                + (f'<div><span style="{BADGE_S}">✓ Seçildi</span></div>' if is_pw else "")
                + '</div>',
                unsafe_allow_html=True,
            )
            if not is_pw:
                st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
                if st.button("Bu Önceliği Seç", key="sel_pw", use_container_width=True):
                    st.session_state.priority_mode = "paper_waste"
                    st.session_state.lambda_val = 1.0
                    st.rerun()

        # ── Kart 2: Zaman Kaybını Azalt ─────────────────────────────────────
        with col2:
            is_tw = (mode == "time_waste")
            st.markdown(
                f'<div style="{_card_wrap(is_tw)}">'
                '<div style="font-size:1.9rem">⏱️</div>'
                f'<div style="{TITLE_S}">Zaman Kaybını Azalt</div>'
                f'<div style="{DESC_S}">Sistem, bobinler arasındaki bıçak düzeni değişimlerini '
                'azaltmaya odaklanır. Üretim akışını hızlandırmak ve hazırlık kayıplarını '
                'azaltmak için uygundur.</div>'
                + (f'<div><span style="{BADGE_S}">✓ Seçildi</span></div>' if is_tw else "")
                + '</div>',
                unsafe_allow_html=True,
            )
            if not is_tw:
                st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
                if st.button("Bu Önceliği Seç", key="sel_tw", use_container_width=True):
                    st.session_state.priority_mode = "time_waste"
                    st.session_state.lambda_val = 0.0
                    st.rerun()

        # ── Kart 3: Dengeli Çözüm ────────────────────────────────────────────
        with col3:
            is_bal = (mode == "balanced")
            st.markdown(
                f'<div style="{_card_wrap(is_bal)}">'
                '<div style="font-size:1.9rem">⚖️</div>'
                f'<div style="{TITLE_S}">Dengeli Çözüm</div>'
                f'<div style="{DESC_S}">Sistem, kağıt kaybı ve zaman kaybı arasında denge kurar. '
                'İsterseniz iki hedefin ağırlığını aşağıdan manuel olarak ayarlayabilirsiniz.</div>'
                + (f'<div><span style="{BADGE_S}">✓ Seçildi</span></div>' if is_bal else "")
                + '</div>',
                unsafe_allow_html=True,
            )
            if not is_bal:
                st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
                if st.button("Bu Önceliği Seç", key="sel_bal", use_container_width=True):
                    st.session_state.priority_mode = "balanced"
                    if st.session_state.lambda_val in (0.0, 1.0):
                        st.session_state.lambda_val = 0.5
                    st.rerun()

        # ── Dengeli Çözüm — Ağırlık Ayar Paneli (hızlı seçim yok) ──────────
        if mode == "balanced":
            st.markdown("<br>", unsafe_allow_html=True)
            lv = float(st.session_state.lambda_val)
            if lv in (0.0, 1.0):
                lv = 0.5
                st.session_state.lambda_val = 0.5

            st.markdown(
                '<div style="background:#F8F9FA;border:1.5px solid #E5E7EB;border-radius:12px;'
                'padding:1.4rem 1.6rem 1.2rem 1.6rem;">'
                '<div style="font-size:.95rem;font-weight:700;color:#1E3A5F;margin-bottom:6px">'
                '⚙️ Ağırlık Ayarı</div>'
                '<div style="font-size:.85rem;color:#4B5563;line-height:1.6;margin-bottom:14px">'
                'Sistem şu anda iki hedefe eşit önem vermektedir. Dilerseniz ağırlıkları '
                'değiştirerek çözümü kağıt kaybı veya zaman kaybı tarafına yaklaştırabilirsiniz.'
                '</div></div>',
                unsafe_allow_html=True,
            )

            _bal_col, _ = st.columns([2, 1])
            with _bal_col:
                st.markdown(
                    '<div style="font-size:.88rem;font-weight:600;color:#374151;margin-bottom:4px">'
                    'Kağıt Kaybı Ağırlığı</div>',
                    unsafe_allow_html=True,
                )

                def _on_bal_slider():
                    st.session_state.lambda_val = round(st.session_state._bal_slider, 2)

                st.slider(
                    "Kağıt Kaybı Ağırlığı",
                    min_value=0.05, max_value=0.95,
                    value=lv, step=0.05,
                    label_visibility="collapsed",
                    key="_bal_slider",
                    on_change=_on_bal_slider,
                )

                lv2 = float(st.session_state.lambda_val)
                st.markdown(
                    f'<div style="display:flex;gap:12px;margin-top:10px">'
                    f'<div style="flex:1;background:#FFF5F5;border-radius:8px;padding:10px;'
                    f'text-align:center;border:1px solid #FECACA">'
                    f'<div style="font-size:1.35rem;font-weight:700;color:#C8102E">{lv2:.0%}</div>'
                    f'<div style="font-size:.75rem;color:#6B7280;margin-top:2px">Kağıt Kaybı Ağırlığı</div>'
                    f'</div>'
                    f'<div style="flex:1;background:#F0FDF4;border-radius:8px;padding:10px;'
                    f'text-align:center;border:1px solid #A7F3D0">'
                    f'<div style="font-size:1.35rem;font-weight:700;color:#059669">{(1-lv2):.0%}</div>'
                    f'<div style="font-size:.75rem;color:#6B7280;margin-top:2px">Zaman Kaybı Ağırlığı</div>'
                    f'</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        # ── Seçim yapılmadan ilerleme engeli ─────────────────────────────────
        st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)
        _nav_cols = st.columns([1, 3, 1])
        with _nav_cols[0]:
            if st.button("← Geri: Sipariş Girişi", use_container_width=True, key="opt_2_back"):
                st.session_state.optimization_step = 1
                st.rerun()
        with _nav_cols[2]:
            if st.button("Kaydet ve Devam Et →", type="primary",
                         use_container_width=True, key="opt_2_next"):
                if mode is None:
                    st.warning("⚠️ Lütfen devam etmeden önce bir öncelik tercihi seçin.")
                else:
                    st.session_state.optimization_step = 3
                    st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    elif step == 3:
        # Step 3 — Kesim Planı Stratejisi
        # ══════════════════════════════════════════════════════════════════════
        _opt_progress_bar(3)
        section_header("🎯", "3. Adım: Kesim Planı Stratejisi",
                       "Üretim planlamasını nasıl yapacağınızı seçin")

        if st.session_state.orders_df is None:
            st.warning("Sipariş verisi bulunamadı. Lütfen 1. Adıma dönün.")
            if st.button("← 1. Adım: Sipariş Girişi", key="opt_3_no_orders"):
                st.session_state.optimization_step = 1; st.rerun()
            return

        df = st.session_state.orders_df

        # ── Helper: _validate_schedule ──────────────────────────────────────
        def _validate_schedule(sdf, valid_types):
            """Returns (ok: bool, errors: list[str])."""
            errs = []
            if sdf is None or sdf.empty:
                errs.append("Bobin sırası tablosu boş. En az bir satır ekleyin.")
                return False, errs
            for idx, row in sdf.iterrows():
                rn = idx + 1
                try:
                    bn = int(float(row["bobin_no"]))
                    if bn <= 0:
                        errs.append(f"Satır {rn}: Bobin numarası pozitif tam sayı olmalıdır.")
                except (ValueError, TypeError):
                    errs.append(f"Satır {rn}: Bobin numarası sayısal bir değer olmalıdır.")
                pt = str(row.get("paper_type", "")).strip()
                if pt == "" or pt == "nan":
                    errs.append(f"Satır {rn}: Kağıt tipi boş olamaz.")
                elif pt not in valid_types:
                    errs.append(f"Satır {rn}: '{pt}' kağıt tipi sipariş listesinde tanımlı değil "
                                f"(geçerli tipler: {', '.join(sorted(valid_types))}).")
            # duplicate bobin_no check
            try:
                nums = [int(float(r["bobin_no"])) for _, r in sdf.iterrows()
                        if str(r.get("bobin_no","")).strip() not in ("","nan")]
                if len(nums) != len(set(nums)):
                    errs.append("Bobin numaraları benzersiz olmalıdır — tekrar eden numara var.")
            except Exception:
                pass
            return len(errs) == 0, errs

        # ── Strategy card style helpers ─────────────────────────────────────
        def _s_card_wrap(selected: bool) -> str:
            if selected:
                return ("border-radius:12px;padding:1.4rem 1.3rem 1rem 1.3rem;"
                        "box-sizing:border-box;background:#FFF5F5;"
                        "border:2.5px solid #C8102E;border-top:4px solid #C8102E;")
            return ("border-radius:12px;padding:1.4rem 1.3rem 1rem 1.3rem;"
                    "box-sizing:border-box;background:#FFFFFF;border:1.5px solid #E5E7EB;")

        S_TITLE = "font-size:1.05rem;font-weight:700;color:#1E3A5F;margin:10px 0 8px 0"
        S_DESC  = "font-size:.87rem;color:#4B5563;line-height:1.65;margin:0"
        S_BADGE = ("display:inline-block;background:#C8102E;color:white;"
                   "border-radius:6px;padding:4px 14px;font-size:.8rem;font-weight:600;margin-top:10px")

        strategy = st.session_state.get("strategy_mode", None)

        # ── 3 strategy cards ────────────────────────────────────────────────
        sc1, sc2, sc3 = st.columns(3)

        with sc1:
            is_gs = (strategy == "given_schedule")
            st.markdown(
                f'<div style="{_s_card_wrap(is_gs)}">'
                f'<div style="font-size:1.7rem;line-height:1">📋</div>'
                f'<div style="{S_TITLE}">Verilen Üretim Sırasına Göre Planla</div>'
                f'<div style="{S_DESC}">Jumbo bobinlerin üretim sırasını ve kağıt tiplerini '
                f'siz belirleyin. Sistem, bu sıraya sadık kalarak optimum kesim planı oluşturur.</div>'
                + (f'<span style="{S_BADGE}">✓ Seçildi</span>' if is_gs else "")
                + "</div>",
                unsafe_allow_html=True,
            )
            if not is_gs:
                if st.button("Bu Stratejiyi Seç", key="sel_gs", use_container_width=True):
                    st.session_state.strategy_mode = "given_schedule"
                    st.session_state.solver_result = None
                    st.rerun()

        with sc2:
            is_at = (strategy == "auto_type")
            st.markdown(
                f'<div style="{_s_card_wrap(is_at)}">'
                f'<div style="font-size:1.7rem;line-height:1">🔄</div>'
                f'<div style="{S_TITLE}">Mevcut Tip Dağılımına Göre Planla</div>'
                f'<div style="{S_DESC}">Bu yöntemde toplam bobin sayısını ve her kağıt tipinden '
                f'kaç bobin kullanılacağını siz belirlersiniz. Bobinlerin kesim sırası sistem '
                f'tarafından belirlenir.</div>'
                + (f'<span style="{S_BADGE}">✓ Seçildi</span>' if is_at else "")
                + "</div>",
                unsafe_allow_html=True,
            )
            if not is_at:
                if st.button("Bu Stratejiyi Seç", key="sel_at", use_container_width=True):
                    st.session_state.strategy_mode = "auto_type"
                    st.session_state.solver_result = None
                    st.rerun()

        with sc3:
            is_fa = (strategy == "full_auto")
            st.markdown(
                f'<div style="{_s_card_wrap(is_fa)}">'
                f'<div style="font-size:1.7rem;line-height:1">⚡</div>'
                f'<div style="{S_TITLE}">Tam Otomatik Plan</div>'
                f'<div style="{S_DESC}">Bu yöntemde kullanıcı yalnızca sipariş bilgilerini girer. '
                f'Sistem gerekli bobin tiplerini, bobin sayısını, kesim sırasını ve kesim desenlerini '
                f'otomatik olarak belirler.</div>'
                + (f'<span style="{S_BADGE}">✓ Seçildi</span>' if is_fa else "")
                + "</div>",
                unsafe_allow_html=True,
            )
            if not is_fa:
                if st.button("Bu Stratejiyi Seç", key="sel_fa", use_container_width=True):
                    st.session_state.strategy_mode = "full_auto"
                    st.session_state.solver_result = None
                    st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Strategy 1: Verilen Üretim Sırasına Göre Planla ─────────────────
        if strategy == "given_schedule":
            unique_types = sorted(df["paper_type"].astype(str).unique())
            type_int_map = {t: i + 1 for i, t in enumerate(unique_types)}

            st.markdown(
                '<div class="info-banner">ℹ️ <b>Nasıl çalışır?</b> Aşağıdaki tabloya jumbo bobinlerinizin '
                'üretim sırasını ve her bobinin kağıt tipini girin. Sistem, bu sıraya sadık kalarak '
                'her bobine hangi siparişlerin kesileceğini belirler.<br>'
                '<b>Not:</b> Kağıt tipi değerleri sipariş listenizdeki tiplerle '
                f'eşleşmelidir ({", ".join(unique_types)}).</div>',
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            # ── Parameters expander ─────────────────────────────────────────
            with st.expander("🔧 Üretim Parametreleri", expanded=False):
                _p1, _p2 = st.columns(2)
                with _p1:
                    st.session_state.L = st.number_input(
                        "Jumbo Bobin Uzunluğu — L (cm)",
                        min_value=1.0, value=float(st.session_state.L),
                        step=1.0, format="%.1f", key="s3_L",
                    )
                    st.session_state.time_limit = st.number_input(
                        "Zaman Limiti (saniye)",
                        min_value=10, max_value=86400,
                        value=int(st.session_state.time_limit),
                        step=60, key="s3_tl",
                    )
                with _p2:
                    st.session_state.B = st.number_input(
                        "Maksimum Bıçak Sayısı — B",
                        min_value=1, max_value=50, value=int(st.session_state.B),
                        step=1, key="s3_B",
                    )

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("#### 📦 Bobin Geliş Sırası")
            st.caption("Üretimde kullanılacak jumbo bobinleri sırasıyla girin.")

            # ── Tabs: Manuel / Excel ────────────────────────────────────────
            tab_man, tab_xls = st.tabs(["✏️  Manuel Giriş", "📁  Excel Yükleme"])

            with tab_man:
                # ── Örnek bobin sırası yükleme ──────────────────────────────
                _SAMPLE_SCHED = [
                    (1, "1.HAMUR OFSET"),
                    (2, "FSC MIX CREDIT KRAFT ÇANTA OB"),
                    (3, "1.HAMUR OFSET"),
                    (4, "FSC MIX CREDIT KRAFT ÇANTA OB"),
                    (5, "1.HAMUR OFSET"),
                    (6, "FSC MIX CREDIT KRAFT ÇANTA OB"),
                    (7, "1.HAMUR OFSET"),
                ]
                if st.button("📋  Örnek Bobin Sırasını Yükle", key="load_sample_sched",
                             use_container_width=False):
                    _ss_df = pd.DataFrame(_SAMPLE_SCHED, columns=["bobin_no", "paper_type"])
                    _ss_df["bobin_no"]   = _ss_df["bobin_no"].astype("Int64")
                    _ss_df["paper_type"] = pt_normalize_series(_ss_df["paper_type"])
                    st.session_state.given_schedule_df = _ss_df
                    st.rerun()

                _sched_init = st.session_state.get("given_schedule_df", None)
                if _sched_init is None or _sched_init.empty:
                    _sched_init = pd.DataFrame({
                        "bobin_no":   pd.Series([], dtype="Int64"),
                        "paper_type": pd.Series([], dtype=str),
                    })
                else:
                    _sched_init = _sched_init.copy()
                    if "bobin_no" not in _sched_init.columns:
                        _sched_init["bobin_no"] = pd.array([], dtype="Int64")
                    if "paper_type" not in _sched_init.columns:
                        _sched_init["paper_type"] = pd.Series([], dtype=str)

                edited_sched = st.data_editor(
                    _sched_init[["bobin_no", "paper_type"]],
                    num_rows="dynamic",
                    use_container_width=True,
                    column_config={
                        "bobin_no":   st.column_config.NumberColumn(
                            "Bobin Numarası", min_value=1, step=1, format="%d",
                            help="Her bobine benzersiz bir numara verin (1, 2, 3…)",
                        ),
                        "paper_type": st.column_config.SelectboxColumn(
                            "Kağıt Tipi",
                            options=unique_types,
                            required=True,
                        ),
                    },
                    key="sched_editor",
                )

                if st.button("💾  Sıralamayı Kaydet", key="save_sched_manual",
                             use_container_width=False):
                    _save = edited_sched.dropna(how="all").copy()
                    _save["paper_type"] = pt_normalize_series(_save["paper_type"].astype(str).str.strip())
                    _save["bobin_no"]   = pd.to_numeric(_save["bobin_no"], errors="coerce").astype("Int64")
                    st.session_state.given_schedule_df = _save
                    st.success("✅ Bobin sırası kaydedildi.")
                    st.rerun()

            with tab_xls:
                # Template download
                _tpl_sched = pd.DataFrame({"Bobin Numarası": [1, 2, 3], "Kağıt Tipi": unique_types[:3] if len(unique_types) >= 3 else (unique_types * 3)[:3]})
                _tpl_buf = io.BytesIO()
                with pd.ExcelWriter(_tpl_buf, engine="openpyxl") as _w:
                    _tpl_sched.to_excel(_w, index=False, sheet_name="Bobin Sırası")
                    _tws_s = _w.sheets["Bobin Sırası"]
                    _tws_s.column_dimensions["B"].width = 36
                    from openpyxl.worksheet.datavalidation import DataValidation as _DVS
                    _dv_formula_s = '"' + ",".join(PAPER_TYPE_OPTIONS) + '"'
                    _dv_s = _DVS(type="list", formula1=_dv_formula_s, sqref="B2:B1000",
                                 showDropDown=False, showErrorMessage=True,
                                 error="Lütfen listeden bir kağıt tipi seçin.",
                                 errorTitle="Geçersiz Kağıt Tipi")
                    _tws_s.add_data_validation(_dv_s)
                _tpl_buf.seek(0)
                st.download_button(
                    "⬇️  Şablon İndir (Excel)",
                    data=_tpl_buf,
                    file_name="bobin_sirasi_sablonu.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_sched_tpl",
                )
                st.caption("Şablonu indirin, doldurun ve yükleyin. Sütun başlıkları: **Bobin Numarası**, **Kağıt Tipi**")

                uploaded_sched = st.file_uploader(
                    "Bobin sırası dosyasını yükleyin (Excel veya CSV)",
                    type=["xlsx", "xls", "csv"],
                    key="sched_uploader",
                )
                if uploaded_sched is not None:
                    try:
                        if uploaded_sched.name.endswith(".csv"):
                            _udf = pd.read_csv(uploaded_sched)
                        else:
                            _udf = pd.read_excel(uploaded_sched)
                        # Normalize column names: Turkish → internal
                        _sched_col_map = {
                            "bobin numarası": "bobin_no", "bobin_numarası": "bobin_no",
                            "bobin no": "bobin_no", "bobin_no": "bobin_no",
                            "kağıt tipi": "paper_type", "kagit tipi": "paper_type",
                            "paper_type": "paper_type", "paper type": "paper_type",
                        }
                        _udf.columns = [_sched_col_map.get(c.strip().lower(), c.strip().lower())
                                        for c in _udf.columns]
                        if "bobin_no" not in _udf.columns or "paper_type" not in _udf.columns:
                            st.error("Dosyada 'Bobin Numarası' ve 'Kağıt Tipi' sütunları bulunamadı.")
                        else:
                            _udf = _udf[["bobin_no", "paper_type"]].dropna(how="all").copy()
                            _udf["paper_type"] = pt_normalize_series(_udf["paper_type"].astype(str).str.strip())
                            _udf["bobin_no"]   = pd.to_numeric(_udf["bobin_no"], errors="coerce").astype("Int64")
                            _unk_s = [v for v in _udf["paper_type"].unique()
                                      if str(v).strip() not in PAPER_TYPE_IDS and str(v).strip() != "nan"]
                            if _unk_s:
                                for _u_s in _unk_s:
                                    st.error(f"Geçersiz kağıt tipi: '{_u_s}'. Lütfen şablondaki kağıt tiplerinden birini seçin.")
                            st.markdown("**Yüklenen tablo önizlemesi:**")
                            st.dataframe(
                                _udf,
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "bobin_no":   st.column_config.NumberColumn("Bobin Numarası"),
                                    "paper_type": st.column_config.TextColumn("Kağıt Tipi"),
                                },
                            )
                            if st.button("✅  Bu Sıralamayı Kullan", key="use_uploaded_sched",
                                         use_container_width=False):
                                st.session_state.given_schedule_df = _udf
                                st.success("Bobin sırası yüklendi.")
                                st.rerun()
                    except Exception as _e:
                        st.error(f"Dosya okunamadı: {_e}")

            # ── Current schedule preview ────────────────────────────────────
            _cur_sched = st.session_state.get("given_schedule_df", None)
            if _cur_sched is not None and not _cur_sched.empty:
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("##### 📋 Kaydedilmiş Bobin Sırası")
                _ok_s, _errs_s = _validate_schedule(_cur_sched, set(unique_types))
                if _ok_s:
                    n_sched = len(_cur_sched)
                    auto_method = "Model 1" if n_sched <= 10 else "Heuristic Model 1"
                    st.markdown(
                        f'<div class="info-banner">🤖 <b>Otomatik Yöntem Seçimi:</b> '
                        f'{n_sched} bobin için <b>{auto_method}</b> kullanılacak.</div>',
                        unsafe_allow_html=True,
                    )
                    st.dataframe(
                        _cur_sched,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "bobin_no":   st.column_config.NumberColumn("Bobin Numarası"),
                            "paper_type": st.column_config.TextColumn("Kağıt Tipi"),
                        },
                    )
                else:
                    for _e in _errs_s:
                        st.error(_e)

            st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)

            # ── Validate params ─────────────────────────────────────────────
            ok_p, errs_p = validate_params(
                st.session_state.L, st.session_state.B,
                st.session_state.CT, st.session_state.lambda_val,
                st.session_state.time_limit,
            )
            if not ok_p:
                for e in errs_p:
                    st.error(e)

            _cur_sched2 = st.session_state.get("given_schedule_df", None)
            _sched_ok, _sched_errs = _validate_schedule(_cur_sched2, set(unique_types))

            _run_disabled = (not ok_p) or (not _sched_ok)

            col_back3, col_run3, col_info3 = st.columns([1, 1, 2])
            with col_back3:
                if st.button("← Öncelik Tercihi", use_container_width=True, key="opt_3_back"):
                    st.session_state.optimization_step = 2; st.rerun()
            with col_run3:
                run_clicked = st.button(
                    "🚀  Optimizasyonu Başlat", type="primary",
                    use_container_width=True, key="opt_run_btn",
                    disabled=_run_disabled,
                )
            with col_info3:
                _n_sc = len(_cur_sched2) if (_cur_sched2 is not None and not _cur_sched2.empty) else 0
                _am   = "Model 1" if _n_sc <= 10 else "Heuristic Model 1"
                st.markdown(
                    f'<div class="info-banner" style="margin:0">'
                    f'Yöntem: <b>{_am}</b> &nbsp;|&nbsp; '
                    f'λ = <b>{st.session_state.lambda_val:.2f}</b> &nbsp;|&nbsp; '
                    f'L = <b>{fmt_n(st.session_state.L)}</b> cm &nbsp;|&nbsp; '
                    f'B = <b>{st.session_state.B}</b> bıçak</div>',
                    unsafe_allow_html=True,
                )

            if run_clicked and _sched_ok and ok_p:
                # Build tj_dict from schedule
                _sched_run = st.session_state.given_schedule_df.copy()
                _sched_run["paper_type"] = _sched_run["paper_type"].astype(str).str.strip()
                _tj_from_sched = {
                    int(row["bobin_no"]): type_int_map[row["paper_type"]]
                    for _, row in _sched_run.iterrows()
                }
                n_sched_run = len(_tj_from_sched)
                st.session_state.method   = "Model 1" if n_sched_run <= 10 else "Heuristic Model 1"
                st.session_state.tj_table = _tj_from_sched
                _run_solver()

            _res = st.session_state.get("solver_result")
            if (_res is not None and _res.get("status") != "error"):
                st.markdown("<br>", unsafe_allow_html=True)
                _cgo, _ = st.columns([1, 3])
                with _cgo:
                    if st.button("📊  Sonuçlara Git →", type="primary",
                                 use_container_width=True, key="opt_goto_results"):
                        st.session_state.active_page = "results"
                        st.rerun()

        # ── Strategy 2: Mevcut Tip Dağılımına Göre Planla ──────────────────
        elif strategy == "auto_type":
            unique_types = sorted(df["paper_type"].astype(str).unique())
            type_int_map = {t: i + 1 for i, t in enumerate(unique_types)}

            st.markdown(
                '<div class="info-banner">ℹ️ <b>Nasıl çalışır?</b> '
                'Bu stratejide hangi kağıt tipinden kaç adet bobin kullanılacağını siz belirlersiniz. '
                'Sistem, bu dağılıma göre bobinlerin kesim sırasını ve kesim desenlerini '
                'otomatik olarak oluşturur.</div>',
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            # ── Parameters expander ─────────────────────────────────────────
            with st.expander("🔧 Üretim Parametreleri", expanded=False):
                _ap1, _ap2 = st.columns(2)
                with _ap1:
                    st.session_state.L = st.number_input(
                        "Jumbo Bobin Uzunluğu — L (cm)",
                        min_value=1.0, value=float(st.session_state.L),
                        step=1.0, format="%.1f", key="at_L",
                    )
                    st.session_state.time_limit = st.number_input(
                        "Zaman Limiti (saniye)",
                        min_value=10, max_value=86400,
                        value=int(st.session_state.time_limit),
                        step=60, key="at_tl",
                    )
                with _ap2:
                    st.session_state.B = st.number_input(
                        "Maksimum Bıçak Sayısı — B",
                        min_value=1, max_value=50, value=int(st.session_state.B),
                        step=1, key="at_B",
                    )

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Bobin Tip Dağılımı input section ───────────────────────────
            st.markdown("#### 📦 Bobin Tip Dağılımı")
            st.caption(
                "Siparişlerde bulunan her kağıt tipi için kaç adet jumbo bobin kullanmak "
                "istediğinizi girin. Toplam bobin sayısı, tip adetlerinin toplamıyla eşleşmelidir."
            )

            # ── Total bobin count input ─────────────────────────────────────
            _at_total = st.number_input(
                "Toplam Jumbo Bobin Sayısı",
                min_value=0,
                value=int(st.session_state.get("_at_user_total", 0)),
                step=1,
                help="Üretimde kullanmak istediğiniz toplam jumbo bobin sayısını girin.",
                key="_at_user_total",
            )

            # ── Per-type count inputs ───────────────────────────────────────
            st.markdown("##### Kağıt Tipi Bazında Bobin Adedi")
            _at_counts = {}
            _at_col_l, _at_col_r = st.columns([1, 1])
            for _i, _tp in enumerate(unique_types):
                _target_col = _at_col_l if _i % 2 == 0 else _at_col_r
                with _target_col:
                    _cnt = st.number_input(
                        f"Kağıt Tipi {_tp} — Bobin Adedi",
                        min_value=0,
                        value=int(st.session_state.get(f"_at_user_cnt_{_tp}", 0)),
                        step=1,
                        key=f"_at_user_cnt_{_tp}",
                    )
                    _at_counts[_tp] = int(_cnt)

            _at_sum = sum(_at_counts.values())

            # ── Validation ──────────────────────────────────────────────────
            _at_valid = True
            _at_errs = []

            _all_zero = (int(_at_total) == 0 and all(v == 0 for v in _at_counts.values()))
            if _all_zero:
                _at_valid = False
            else:
                if int(_at_total) <= 0:
                    _at_errs.append("Toplam bobin sayısı pozitif bir tam sayı olmalıdır.")
                    _at_valid = False

                if _at_sum != int(_at_total):
                    _at_errs.append(
                        f"Tip bazlı bobin adetlerinin toplamı ({_at_sum}), "
                        f"toplam bobin sayısıyla ({int(_at_total)}) eşleşmiyor. Lütfen düzeltin."
                    )
                    _at_valid = False

                for _tp in unique_types:
                    if _at_counts.get(_tp, 0) == 0:
                        _at_errs.append(f"Kağıt Tipi '{_tp}' için bobin adedi girilmedi.")
                        _at_valid = False

                if _at_errs:
                    for _ae in _at_errs:
                        st.error(_ae)
                elif _at_valid:
                    _at_method_label = "Model 2" if int(_at_total) <= 7 else "Heuristic Model 2"
                    st.markdown(
                        f'<div class="info-banner">🤖 <b>Otomatik Yöntem Seçimi:</b> '
                        f'{int(_at_total)} bobin için <b>{_at_method_label}</b> kullanılacak.</div>',
                        unsafe_allow_html=True,
                    )

            # ── Validate params ─────────────────────────────────────────────
            ok_p_at, errs_p_at = validate_params(
                st.session_state.L, st.session_state.B,
                st.session_state.CT, st.session_state.lambda_val,
                st.session_state.time_limit,
            )
            if not ok_p_at:
                for _e in errs_p_at:
                    st.error(_e)

            _run_disabled_at = (not _at_valid) or (not ok_p_at) or _all_zero

            st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)

            col_back_at, col_run_at, col_info_at = st.columns([1, 1, 2])
            with col_back_at:
                if st.button("← Öncelik Tercihi", use_container_width=True, key="opt_3_back"):
                    st.session_state.optimization_step = 2; st.rerun()
            with col_run_at:
                run_clicked_at = st.button(
                    "🚀  Optimizasyonu Başlat", type="primary",
                    use_container_width=True, key="opt_run_btn",
                    disabled=_run_disabled_at,
                )
            with col_info_at:
                if not _all_zero and _at_valid:
                    _at_ml = "Model 2" if int(_at_total) <= 7 else "Heuristic Model 2"
                    st.markdown(
                        f'<div class="info-banner" style="margin:0">'
                        f'Yöntem: <b>{_at_ml}</b> &nbsp;|&nbsp; '
                        f'λ = <b>{st.session_state.lambda_val:.2f}</b> &nbsp;|&nbsp; '
                        f'L = <b>{fmt_n(st.session_state.L)}</b> cm &nbsp;|&nbsp; '
                        f'B = <b>{st.session_state.B}</b> bıçak</div>',
                        unsafe_allow_html=True,
                    )

            if run_clicked_at and _at_valid and ok_p_at and not _all_zero:
                # Build TJ from user's per-type counts (sequential blocks by type)
                _tj_at = {}
                _s_idx = 1
                for _tp in unique_types:
                    for _ in range(_at_counts.get(_tp, 0)):
                        _tj_at[_s_idx] = type_int_map[_tp]
                        _s_idx += 1

                _n_at = int(_at_total)
                st.session_state.method   = "Model 2" if _n_at <= 7 else "Heuristic Model 2"
                st.session_state.tj_table = _tj_at
                _run_solver()

            # ── Post-solve banners ──────────────────────────────────────────
            _res_at = st.session_state.get("solver_result")

            if _res_at is not None and _res_at.get("status") == "error":
                st.markdown("<br>", unsafe_allow_html=True)
                st.error(
                    "**Çözüm bulunamadı.** Girilen bobin dağılımı siparişleri karşılamak için "
                    "yeterli olmayabilir. Lütfen bobin adetlerini artırarak tekrar deneyin."
                )

            if (_res_at is not None and _res_at.get("status") != "error"):
                st.markdown("<br>", unsafe_allow_html=True)
                _cgo2, _ = st.columns([1, 3])
                with _cgo2:
                    if st.button("📊  Sonuçlara Git →", type="primary",
                                 use_container_width=True, key="opt_goto_results"):
                        st.session_state.active_page = "results"
                        st.rerun()

        # ── Strategy 3: Tam Otomatik Plan ───────────────────────────────────
        elif strategy == "full_auto":

            # ── Info panel — no extra inputs ────────────────────────────────
            st.markdown(
                '<div style="background:#F0F4FF;border:1.5px solid #C7D7F5;border-radius:10px;'
                'padding:1.1rem 1.3rem;">'
                '<div style="font-weight:700;color:#1E3A5F;font-size:1rem;margin-bottom:6px">'
                '⚡ Tam Otomatik Planlama</div>'
                '<div style="font-size:.88rem;color:#374151;line-height:1.65;">'
                'Bu stratejide sistem, girilen siparişleri kullanarak gerekli bobin tiplerini, '
                'bobin sayılarını, kesim sırasını ve kesim desenlerini otomatik olarak belirler. '
                'Kullanıcıdan ek bir bobin bilgisi alınmaz.</div>'
                '<div style="font-size:.82rem;color:#6B7280;margin-top:8px;font-style:italic;">'
                'Bu seçenek, bobin dağılımı ve üretim sırası önceden bilinmediğinde '
                'kullanılabilir.</div>'
                '</div>',
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

            # ── Parameters expander ─────────────────────────────────────────
            with st.expander("🔧 Üretim Parametreleri", expanded=False):
                _fp1, _fp2 = st.columns(2)
                with _fp1:
                    st.session_state.L = st.number_input(
                        "Jumbo Bobin Uzunluğu — L (cm)",
                        min_value=1.0, value=float(st.session_state.L),
                        step=1.0, format="%.1f", key="fa_L",
                    )
                    st.session_state.time_limit = st.number_input(
                        "Zaman Limiti (saniye)",
                        min_value=10, max_value=86400,
                        value=int(st.session_state.time_limit),
                        step=60, key="fa_tl",
                    )
                with _fp2:
                    st.session_state.B = st.number_input(
                        "Maksimum Bıçak Sayısı — B",
                        min_value=1, max_value=50, value=int(st.session_state.B),
                        step=1, key="fa_B",
                    )

            # ── Sipariş validation ──────────────────────────────────────────
            ok_ord, errs_ord = validate_orders(df)
            if not ok_ord:
                for _oe in errs_ord:
                    st.error(_oe)

            # Length > L check
            _fa_len_errs = []
            for _idx, _row in df.iterrows():
                try:
                    _rlen = float(_row["length"])
                    if _rlen > float(st.session_state.L):
                        _fa_len_errs.append(
                            f"Satır {_idx + 2}: Sipariş uzunluğu ({_rlen} cm) jumbo bobin "
                            f"uzunluğundan ({int(st.session_state.L)} cm) büyük olamaz."
                        )
                except (ValueError, TypeError):
                    pass
            for _le in _fa_len_errs:
                st.error(_le)

            # ── Auto-method selection banner ────────────────────────────────
            _, _fa_est_min = estimate_reel_count(df, st.session_state.L, st.session_state.B)
            st.markdown(
                '<div class="info-banner">🤖 Sistem, problem büyüklüğüne göre uygun çözüm '
                'yöntemini otomatik olarak seçer.</div>',
                unsafe_allow_html=True,
            )

            # ── Validate params ─────────────────────────────────────────────
            ok_p_fa, errs_p_fa = validate_params(
                st.session_state.L, st.session_state.B,
                st.session_state.CT, st.session_state.lambda_val,
                st.session_state.time_limit,
            )
            if not ok_p_fa:
                for _e in errs_p_fa:
                    st.error(_e)

            _fa_run_disabled = (not ok_ord) or bool(_fa_len_errs) or (not ok_p_fa)

            st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)

            col_back_fa, col_run_fa, col_info_fa = st.columns([1, 1, 2])
            with col_back_fa:
                if st.button("← Öncelik Tercihi", use_container_width=True, key="opt_3_back_ph"):
                    st.session_state.optimization_step = 2; st.rerun()
            with col_run_fa:
                run_clicked_fa = st.button(
                    "🚀  Optimizasyonu Başlat", type="primary",
                    use_container_width=True, key="opt_run_btn",
                    disabled=_fa_run_disabled,
                )
            with col_info_fa:
                st.markdown(
                    f'<div class="info-banner" style="margin:0">'
                    f'λ = <b>{st.session_state.lambda_val:.2f}</b> &nbsp;|&nbsp; '
                    f'L = <b>{fmt_n(st.session_state.L)}</b> cm &nbsp;|&nbsp; '
                    f'B = <b>{st.session_state.B}</b> bıçak</div>',
                    unsafe_allow_html=True,
                )

            if run_clicked_fa and ok_ord and not _fa_len_errs and ok_p_fa:
                # No TJ — system decides types and order (Model 2 / Heuristic Model 2)
                st.session_state.tj_table = None
                st.session_state.method = "Model 2" if _fa_est_min <= 7 else "Heuristic Model 2"
                _run_solver()

            # ── Post-solve banners ──────────────────────────────────────────
            _res_fa = st.session_state.get("solver_result")

            if _res_fa is not None and _res_fa.get("status") == "error":
                st.markdown("<br>", unsafe_allow_html=True)
                st.error(
                    "**Çözüm bulunamadı.** Lütfen sipariş bilgilerini kontrol ederek tekrar deneyin."
                )

            if (_res_fa is not None and _res_fa.get("status") != "error"):
                st.markdown("<br>", unsafe_allow_html=True)
                _cgo3, _ = st.columns([1, 3])
                with _cgo3:
                    if st.button("📊  Sonuçlara Git →", type="primary",
                                 use_container_width=True, key="opt_goto_results"):
                        st.session_state.active_page = "results"
                        st.rerun()

        # ── No strategy selected ─────────────────────────────────────────────
        else:
            st.markdown("<br>", unsafe_allow_html=True)
            st.warning("⚠️ Lütfen devam etmeden önce bir strateji seçin.")
            _nb2, _ = st.columns([1, 3])
            with _nb2:
                if st.button("← Öncelik Tercihi", use_container_width=True, key="opt_3_back_ns"):
                    st.session_state.optimization_step = 2; st.rerun()


# ===========================================================================
# PAGE: SİPARİŞ GİRİŞİ
# ===========================================================================
def page_orders():
    section_header("📋", "Sipariş Girişi", "Müşteri siparişlerini manuel girin veya dosya yükleyin")

    tab_manual, tab_upload = st.tabs(["✏️  Manuel Giriş", "📁  Excel / CSV Yükleme"])

    # ── Manual entry ───────────────────────────────────────────────────────
    with tab_manual:
        st.markdown("##### Sipariş Tablosu")
        st.caption("Satır eklemek için son satırın altına tıklayın. Kağıt tipi olarak rakam veya kod girebilirsiniz.")

        if st.session_state.orders_df is not None:
            init_df = normalize_orders_df(st.session_state.orders_df)
            for col in ["paper_type", "length", "quantity"]:
                if col not in init_df.columns:
                    init_df[col] = ""
        else:
            # Start empty — user must explicitly load sample or type in rows
            init_df = pd.DataFrame({
                "paper_type": pd.Series([], dtype=str),
                "length":     pd.Series([], dtype=float),
                "quantity":   pd.array([], dtype="Int64"),
            })

        edited = st.data_editor(
            init_df[["paper_type", "length", "quantity"]],
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "paper_type": st.column_config.SelectboxColumn(
                    "Kağıt Tipi", options=PAPER_TYPE_OPTIONS, required=True,
                ),
                "length":     st.column_config.NumberColumn("Uzunluk (cm)", min_value=0.1, format="%.1f"),
                "quantity":   st.column_config.NumberColumn("Miktar (adet)", min_value=1, step=1, format="%d"),
            },
            hide_index=True,
            key="orders_editor",
        )

        col_a, col_b, col_c = st.columns([1, 1, 2])
        with col_a:
            if st.button("✅  Siparişleri Kaydet", type="primary", use_container_width=True):
                df = normalize_orders_df(edited.dropna(how="all").reset_index(drop=True))
                df["paper_type"] = pt_normalize_series(df["paper_type"])
                ok, errs = validate_orders(df)
                if ok:
                    st.session_state.orders_df = df
                    st.session_state.solver_result = None
                    st.success(f"✓ {len(df)} sipariş kaydedildi.")
                else:
                    for e in errs:
                        st.error(e)
        with col_b:
            if st.button("🔄  Örnek Veriyi Yükle", use_container_width=True):
                _csv_path = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "example_orders.csv",
                )
                _s = get_sample_orders_df(_csv_path)
                _s["paper_type"] = pt_normalize_series(_s["paper_type"])
                st.session_state.orders_df = _s
                st.session_state.solver_result = None
                st.rerun()
        with col_c:
            try:
                import io as _tio
                from openpyxl import Workbook as _WB
                from openpyxl.styles import Font as _Fnt, PatternFill as _PF, Alignment as _Al
                from openpyxl.worksheet.datavalidation import DataValidation as _DV
                _tbuf = _tio.BytesIO()
                _twb  = _WB()
                _tws  = _twb.active
                _tws.title = "Siparişler"
                _red_fill = _PF(start_color="C8102E", end_color="C8102E", fill_type="solid")
                _hdr_fnt  = _Fnt(bold=True, color="FFFFFF", size=11)
                for _ci, _h in enumerate(["Kağıt Tipi", "Uzunluk", "Miktar"], start=1):
                    _c = _tws.cell(row=1, column=_ci, value=_h)
                    _c.fill = _red_fill; _c.font = _hdr_fnt
                    _c.alignment = _Al(horizontal="center")
                _ex_rows = [
                    ("1.HAMUR OFSET", 140.0, 5),
                    ("1.HAMUR IVORY(D)", 170.0, 4),
                ]
                for _ri, (_pt, _ln, _qty) in enumerate(_ex_rows, start=2):
                    _tws.cell(row=_ri, column=1, value=_pt)
                    _tws.cell(row=_ri, column=2, value=_ln)
                    _tws.cell(row=_ri, column=3, value=_qty)
                # Bilgi notu — tablonun sağına (E1:G3), açık sarı arka plan
                from openpyxl.styles import PatternFill as _NFill, Font as _NFont, Alignment as _NAl
                _note_txt = (
                    "⚠  Bu satırlar yalnızca örnek veri formatını göstermek için eklenmiştir.\n"
                    "Kendi siparişlerinizi girmeden önce bu örnek satırları silin\n"
                    "ve yerine gerçek verilerinizi yazın."
                )
                _tws.merge_cells("E1:G3")
                _nc = _tws["E1"]
                _nc.value = _note_txt
                _nc.fill = _NFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
                _nc.font = _NFont(size=9, italic=True, color="5C4A00")
                _nc.alignment = _NAl(wrap_text=True, vertical="top", horizontal="left")
                _dv_f = '"' + ",".join(PAPER_TYPE_OPTIONS) + '"'
                _dv = _DV(type="list", formula1=_dv_f, sqref="A2:A1000",
                          showDropDown=False, showErrorMessage=True,
                          error="Lütfen listeden bir kağıt tipi seçin.",
                          errorTitle="Geçersiz Kağıt Tipi")
                _tws.add_data_validation(_dv)
                from openpyxl.utils import get_column_letter as _gcl
                for _ci, _w in [(1, 36), (2, 14), (3, 14), (4, 4), (5, 18), (6, 18), (7, 18)]:
                    _tws.column_dimensions[_gcl(_ci)].width = _w
                _twb.save(_tbuf); _tbuf.seek(0)
                st.download_button(
                    "⬇️  Boş Şablon İndir (.xlsx)",
                    data=_tbuf,
                    file_name="siparis_sablonu.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception:
                _tcsv = (
                    "Kağıt Tipi,Uzunluk,Miktar\n"
                    "1.HAMUR OFSET,140,5\n"
                    "1.HAMUR IVORY(D),170,4\n"
                )
                st.download_button(
                    "⬇️  Boş Şablon İndir (.csv)",
                    data=_tcsv,
                    file_name="siparis_sablonu.csv",
                    mime="text/csv",
                    use_container_width=True,
                )


    # ── Upload ─────────────────────────────────────────────────────────────
    with tab_upload:
        st.markdown("##### Dosya Yükleme")
        st.caption("CSV veya Excel (xlsx) dosyası yükleyin. Gerekli sütunlar: **Kağıt Tipi**, **Uzunluk**, **Miktar**")

        uploaded = st.file_uploader(
            "Dosya seçin",
            type=["csv", "xlsx", "xls"],
            label_visibility="collapsed",
        )

        if uploaded:
            try:
                if uploaded.name.endswith((".xlsx", ".xls")):
                    df_up = pd.read_excel(uploaded)
                else:
                    df_up = pd.read_csv(uploaded)

                # Lowercase column names
                df_up.columns = [c.strip().lower().replace(" ", "_") for c in df_up.columns]

                # Türkçe kolon adlarını iç isimlere çevir
                _tr_to_en = {
                    "kağıt_tipi": "paper_type", "kagit_tipi": "paper_type",
                    "uzunluk": "length", "uzunluk_(cm)": "length",
                    "miktar": "quantity", "miktar_(adet)": "quantity",
                }
                df_up = df_up.rename(columns={k: v for k, v in _tr_to_en.items() if k in df_up.columns})

                # İsteğe bağlı sipariş/order numarası sütununu kaldır
                for _drop_col in ["order_id", "sipariş_numarası", "siparis_numarasi"]:
                    if _drop_col in df_up.columns:
                        df_up = df_up.drop(columns=[_drop_col])

                df_up = normalize_orders_df(df_up.reset_index(drop=True))

                st.markdown("**Yüklenen veri önizlemesi:**")
                st.dataframe(df_up.head(20), use_container_width=True, hide_index=True,
                    column_config={
                        "paper_type": st.column_config.TextColumn("Kağıt Tipi"),
                        "length":     st.column_config.NumberColumn("Uzunluk (cm)"),
                        "quantity":   st.column_config.NumberColumn("Miktar (adet)"),
                    })

                ok, errs = validate_orders(df_up)
                if ok:
                    st.markdown(
                        f'<div class="info-banner">✓ {len(df_up)} sipariş doğrulandı ve hazır.</div>',
                        unsafe_allow_html=True,
                    )
                    if st.button("✅  Bu Veriyi Kullan", type="primary"):
                        st.session_state.orders_df = df_up
                        st.session_state.solver_result = None
                        st.success("Sipariş verisi kaydedildi.")
                else:
                    for e in errs:
                        st.error(e)
            except Exception as e:
                st.error(f"Dosya okunurken hata: {e}")


    # ── Current orders preview ─────────────────────────────────────────────
    if st.session_state.orders_df is not None:
        df = st.session_state.orders_df
        st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)
        st.markdown("#### 📌 Mevcut Siparişler")
        c1, c2 = st.columns(2)
        with c1:
            metric_card(len(df), "Toplam Sipariş Satırı")
        with c2:
            metric_card(len(df["paper_type"].unique()), "Farklı Kağıt Tipi")

        with st.expander("Sipariş tablosunu göster"):
            st.dataframe(df, use_container_width=True, hide_index=True,
                column_config={
                    "paper_type": st.column_config.TextColumn("Kağıt Tipi"),
                    "length":     st.column_config.NumberColumn("Uzunluk (cm)"),
                    "quantity":   st.column_config.NumberColumn("Miktar (adet)"),
                })

        col_next, _ = st.columns([1, 3])
        with col_next:
            if st.button("⚙️  Çözüm Ayarlarına Geç →", type="primary", use_container_width=True):
                nav_to("settings")
                st.rerun()


# ===========================================================================
# PAGE: ÇÖZÜM AYARLARI
# ===========================================================================
def page_settings():
    section_header("⚙️", "Çözüm Ayarları", "Yöntem seçimi, lambda ve parametre ayarları")

    if st.session_state.orders_df is None:
        st.markdown(
            '<div class="warn-banner">⚠️ Henüz sipariş girilmedi. Önce Sipariş Girişi sayfasını tamamlayın.</div>',
            unsafe_allow_html=True,
        )
        if st.button("📋  Sipariş Girişine Git"):
            st.session_state.optimization_step = 1; nav_to("optimization"); st.rerun()
        return

    df = st.session_state.orders_df

    # ── Reel count estimation & warning ────────────────────────────────────
    safe_count, est_min = estimate_reel_count(df, st.session_state.L, st.session_state.B)
    st.session_state.safe_count = safe_count

    col_left, col_right = st.columns([1.2, 1])

    # ── Method selection ───────────────────────────────────────────────────
    with col_left:
        st.markdown("#### Çözüm Yöntemi")
        
        # Determine recommendation dynamically based on est_min (estimated required reels)
        rec_model2 = (est_min <= 10)
        rec_heuristic2 = (est_min > 10)
        
        # Auto-select recommended method when est_min changes across threshold, or first load
        recommended_method = "Heuristic Model 2" if est_min > 10 else "Model 2"
        if st.session_state.get("_prev_est_min") != est_min:
            prev = st.session_state.get("_prev_est_min")
            if prev is None or (prev <= 10 and est_min > 10) or (prev > 10 and est_min <= 10):
                st.session_state.method = recommended_method
            st.session_state._prev_est_min = est_min

        method_opts = {
            "Model 1": {
                "desc": "Kesin model — Bobin tipleri önceden belirlenir. Gurobi gerektirir.",
                "rec": False,
            },
            "Model 2": {
                "desc": "Kesin model — Bobin tipini otomatik seçer. Gurobi gerektirir.",
                "rec": rec_model2,
            },
            "Heuristic Model 1": {
                "desc": "Hızlı sezgisel — Bobin tipleri önceden belirlenir. Lisans gerektirmez.",
                "rec": False,
            },
            "Heuristic Model 2": {
                "desc": "Hızlı sezgisel — Bobin tipini otomatik seçer. Lisans gerektirmez.",
                "rec": rec_heuristic2,
            },
        }

        # Show currently selected method prominently
        sel_method = st.session_state.method
        sel_info = method_opts.get(sel_method, method_opts["Heuristic Model 2"])
        
        # Warnings for exact models
        if sel_method in ["Model 1", "Model 2"]:
            st.markdown(
                '<div class="warn-banner">⚠️ <b>Uyarı:</b> Bu yöntem Gurobi lisansı gerektirir. '
                'Bulut ortamında çalışması için Gurobi WLS lisansı tanımlanmalıdır.</div>',
                unsafe_allow_html=True,
            )
            if est_min > 10:
                st.markdown(
                    '<div class="warn-banner">⚠️ <b>Büyük Veri Seti:</b> Bu veri seti büyük görünüyor. '
                    'Exact modeller uzun sürebilir. Daha hızlı sonuç için <b>Heuristic Model 2</b> önerilir.</div>',
                    unsafe_allow_html=True,
                )

        # Selected solver card
        prefix = "Önerilen Çözüm Yöntemi" if sel_info["rec"] else "Seçili Çözüm Yöntemi"
        st.markdown(
            f'''<div class="method-card selected" style="margin-bottom:1rem; padding: 20px;">
                <span class="method-title" style="font-size:1.1rem;">{prefix}: {sel_method}</span>
                <div class="method-desc" style="font-size:0.9rem; margin-top:8px;">{sel_info["desc"]}</div>
            </div>''',
            unsafe_allow_html=True,
        )

        with st.expander("⚙️ Ekstra Çözüm Yöntemleri"):
            st.caption("Farklı bir algoritma seçmek için aşağıdaki yöntemlerden birini kullanabilirsiniz.")
            
            st.markdown("##### Verilen Üretim Sırasına Göre Çözüm")
            st.caption("Bu yöntemlerde jumbo bobin tipleri önceden kullanıcı tarafından belirlenir.")
            for mname in ["Model 1", "Heuristic Model 1"]:
                minfo = method_opts[mname]
                is_sel = st.session_state.method == mname
                rec_badge = '<span class="recommended-badge">ÖNERİLEN</span>' if minfo["rec"] else ""
                border = "border: 2px solid #C8102E; background: #FFF5F5;" if is_sel else "border: 2px solid #E5E7EB;"
                st.markdown(
                    f'''<div class="method-card" style="{border}">
                            <span class="method-title">{mname}</span>{rec_badge}
                            <div class="method-desc">{minfo["desc"]}</div>
                        </div>''',
                    unsafe_allow_html=True,
                )
                if st.button(f"Seç: {mname}", key=f"sel_{mname}", use_container_width=True, type="primary" if is_sel else "secondary"):
                    st.session_state.method = mname; st.session_state.solver_result = None; st.rerun()

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("##### Otomatik Bobin Tipi Seçimli Çözüm")
            st.caption("Bu yöntemlerde jumbo bobin tipleri model tarafından otomatik seçilir.")
            for mname in ["Model 2", "Heuristic Model 2"]:
                minfo = method_opts[mname]
                is_sel = st.session_state.method == mname
                rec_badge = '<span class="recommended-badge">ÖNERİLEN</span>' if minfo["rec"] else ""
                border = "border: 2px solid #C8102E; background: #FFF5F5;" if is_sel else "border: 2px solid #E5E7EB;"
                st.markdown(
                    f'''<div class="method-card" style="{border}">
                            <span class="method-title">{mname}</span>{rec_badge}
                            <div class="method-desc">{minfo["desc"]}</div>
                        </div>''',
                    unsafe_allow_html=True,
                )
                if st.button(f"Seç: {mname}", key=f"sel_{mname}", use_container_width=True, type="primary" if is_sel else "secondary"):
                    st.session_state.method = mname; st.session_state.solver_result = None; st.rerun()


    # ── Lambda & advanced params ───────────────────────────────────────────
    with col_right:
        st.markdown("#### Lambda (λ) Ağırlık Parametresi")
        st.caption("λ → Kağıt Kaybı ağırlığı &nbsp;|&nbsp; (1−λ) → Zaman Kaybı ağırlığı")

        def _on_lambda_slider():
            st.session_state.lambda_val = round(st.session_state._lslider, 2)

        def _on_lambda_num():
            st.session_state.lambda_val = round(st.session_state._lnum, 2)

        st.slider(
            "λ değeri",
            min_value=0.0, max_value=1.0,
            value=float(st.session_state.lambda_val),
            step=0.05,
            label_visibility="collapsed",
            key="_lslider",
            on_change=_on_lambda_slider,
        )
        st.number_input(
            "Tam değer girin (0–1)",
            min_value=0.0, max_value=1.0,
            value=float(st.session_state.lambda_val),
            step=0.01,
            format="%.2f",
            key="_lnum",
            on_change=_on_lambda_num,
        )

        lam_val = st.session_state.lambda_val

        col_lv, col_rv = st.columns(2)
        with col_lv:
            st.markdown(
                f'<div style="text-align:center;background:#FFF5F5;border-radius:8px;padding:10px">'
                f'<div style="font-size:1.4rem;font-weight:700;color:#C8102E">{lam_val:.2f}</div>'
                f'<div style="font-size:.72rem;color:#6B7280">λ — Kağıt Kaybı</div></div>',
                unsafe_allow_html=True,
            )
        with col_rv:
            st.markdown(
                f'<div style="text-align:center;background:#F0FDF4;border-radius:8px;padding:10px">'
                f'<div style="font-size:1.4rem;font-weight:700;color:#059669">{1-lam_val:.2f}</div>'
                f'<div style="font-size:.72rem;color:#6B7280">1−λ — Zaman Kaybı</div></div>',
                unsafe_allow_html=True,
            )

        st.markdown("<br>", unsafe_allow_html=True)

        with st.expander("🔧 Gelişmiş Parametre Ayarları"):
            st.session_state.L = st.number_input(
                "Jumbo Bobin Uzunluğu — L (cm)",
                min_value=1.0, value=float(st.session_state.L), step=1.0, format="%.1f",
            )
            st.session_state.B = st.number_input(
                "Maksimum Bıçak Sayısı — B",
                min_value=1, max_value=50, value=int(st.session_state.B), step=1,
            )
            st.session_state.time_limit = st.number_input(
                "Zaman Limiti — saniye (Model 1 & 2)",
                min_value=10, max_value=86400, value=int(st.session_state.time_limit), step=60,
            )

            # Re-estimate after param change (internal values, not shown to user)
            safe_count, est_min = estimate_reel_count(df, st.session_state.L, st.session_state.B)
            st.session_state.safe_count = safe_count


    # ── Jumbo Bobin Type input (Model 1 / Heuristic 1 only) ────────────────
    method = st.session_state.method
    if method in ("Model 1", "Heuristic Model 1"):
        st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)
        section_header("🔢", "Jumbo Bobin Tip Ataması",
                        "Model 1 / Heuristic Model 1: Hangi bobine hangi kağıt tipi atanacak?")

        unique_types = sorted(df["paper_type"].astype(str).unique())
        type_int_map = {t: i + 1 for i, t in enumerate(unique_types)}

        # ── Per-type and total estimates (user-facing) ─────────────────────
        est_total, per_type_est = estimate_reels_detailed(df, st.session_state.L)

        # Reset counts whenever the underlying estimates change (new orders or L changed)
        _est_hash = (est_total, tuple(sorted(per_type_est.items())))
        if st.session_state.get("_tj_est_hash") != _est_hash:
            st.session_state["_tj_est_hash"] = _est_hash
            st.session_state["user_reel_count"] = est_total
            for tp, cnt in per_type_est.items():
                st.session_state[f"tj_cnt_{tp}"] = cnt

        # ── User-facing info ───────────────────────────────────────────────
        dist_parts = " | ".join(
            f"{tp}: {per_type_est.get(tp, 1)}" for tp in unique_types
        )
        st.info(
            f"Tahmini gerekli bobin sayısı: **{est_total}**\n\n"
            f"Tip bazında önerilen dağılım: {dist_parts}",
            icon="💡",
        )
        st.caption(
            "Bu değerler tahminidir. "
            "Çözüm bulunamazsa ilgili tip için bobin sayısını artırın."
        )

        # ── Total reel count input ─────────────────────────────────────────
        if "user_reel_count" not in st.session_state:
            st.session_state["user_reel_count"] = est_total

        n_reels = st.number_input(
            "Toplam Jumbo Bobin Sayısı",
            min_value=max(1, len(unique_types)),
            value=int(st.session_state["user_reel_count"]),
            step=1,
            help=(
                f"Tahmini minimum: {est_total}. "
                "Çözücü yalnızca burada belirlenen bobinleri kullanır."
            ),
            key="user_reel_count",
        )
        n_reels = int(n_reels)

        col_tj_l, col_tj_r = st.columns(2)
        with col_tj_l:
            st.caption(f"Toplam bobin: {n_reels} | Kağıt tipleri: {', '.join(unique_types)}")

            counts = {}
            remaining_reels = n_reels
            for i, tp in enumerate(unique_types):
                # Default for this type: use per-type estimate (already set in session state)
                _default = per_type_est.get(tp, max(1, n_reels // len(unique_types)))

                if i < len(unique_types) - 1:
                    max_cnt = remaining_reels - (len(unique_types) - i - 1)
                    val = st.number_input(
                        f"{tp} — Bobin Sayısı  (önerilen: {per_type_est.get(tp, '?')})",
                        min_value=1,
                        max_value=max(1, max_cnt),
                        value=min(max(1, _default), max(1, max_cnt)),
                        step=1,
                        key=f"tj_cnt_{tp}",
                    )
                    counts[tp] = val
                    remaining_reels -= val
                else:
                    counts[tp] = max(1, remaining_reels)
                    st.markdown(
                        f'<div style="padding:8px;background:#F3F4F6;border-radius:8px;'
                        f'margin-top:8px">'
                        f"<b>{tp}</b>: <b>{counts[tp]}</b> bobin (kalan) "
                        f"— önerilen: {per_type_est.get(tp, '?')}</div>",
                        unsafe_allow_html=True,
                    )

        with col_tj_r:
            # Build TJ dict exactly from user counts — no auto-fill, no silent extension
            tj = {}
            s_idx = 1
            for tp in unique_types:
                for _ in range(counts.get(tp, 1)):
                    tj[s_idx] = type_int_map[tp]
                    s_idx += 1

            st.session_state.tj_table = tj
            tj_df = pd.DataFrame(
                [(k, v) for k, v in tj.items()],
                columns=["Bobin Numarası", "Tip"],
            )
            st.caption("Oluşturulan atama tablosu (ilk 15 satır)")
            st.dataframe(tj_df.head(15), use_container_width=True, hide_index=True, height=300)


    # ── Run button ─────────────────────────────────────────────────────────
    st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)

    ok_p, errs_p = validate_params(
        st.session_state.L, st.session_state.B,
        st.session_state.CT, st.session_state.lambda_val,
        st.session_state.time_limit,
    )
    if not ok_p:
        for e in errs_p:
            st.error(e)

    col_run, col_info = st.columns([1, 3])
    with col_run:
        run_clicked = st.button(
            "🚀  Optimizasyonu Başlat",
            type="primary",
            use_container_width=True,
            disabled=not ok_p,
        )
    with col_info:
        st.markdown(
            f'<div class="info-banner" style="margin:0">'
            f'Seçili yöntem: <b>{st.session_state.method}</b> &nbsp;|&nbsp; '
            f'λ = <b>{st.session_state.lambda_val:.2f}</b> &nbsp;|&nbsp; '
            f'L = <b>{fmt_n(st.session_state.L)}</b> cm &nbsp;|&nbsp; '
            f'B = <b>{st.session_state.B}</b> bıçak</div>',
            unsafe_allow_html=True,
        )

    if run_clicked:
        _run_solver()

    # Persistent navigation button — visible whenever a result exists
    if st.session_state.solver_result is not None and \
            st.session_state.solver_result.get("status") != "error":
        st.markdown("<br>", unsafe_allow_html=True)
        col_go, _ = st.columns([1, 3])
        with col_go:
            if st.button("📊  Sonuçlara Git →", type="primary",
                         use_container_width=True, key="goto_results_btn"):
                st.session_state.active_page = "results"
                st.rerun()


def _run_solver():
    df = st.session_state.orders_df
    method = st.session_state.method

    ok_p, errs_p = validate_params(
        st.session_state.L, st.session_state.B,
        st.session_state.CT, st.session_state.lambda_val,
        st.session_state.time_limit,
    )
    if not ok_p:
        for e in errs_p:
            st.error(e)
        return

    tj_dict = st.session_state.tj_table if method in ("Model 1", "Heuristic Model 1") else None

    try:
        data, type_map, safe_count = build_data_dict(
            df,
            L=st.session_state.L,
            B=st.session_state.B,
            CT=st.session_state.CT,
            lam=st.session_state.lambda_val,
            method=method,
            tj_dict=tj_dict,
        )
    except Exception as e:
        st.error(f"Veri hazırlanırken hata: {e}")
        return

    st.session_state.data_dict = data
    st.session_state.type_map = type_map

    spinner_msg = "⚙️ Optimizasyon çalışıyor, lütfen bekleyin..."
    if method in ("Model 1", "Model 2"):
        spinner_msg += f" (Zaman limiti: {st.session_state.time_limit} sn)"

    with st.spinner(spinner_msg):
        result = run_selected_solver(
            method_name=method,
            data=data,
            time_limit=int(st.session_state.time_limit),
            verbose=False,
        )

    st.session_state.solver_result = result
    st.session_state["reel_warn_dismissed"] = False  # reset on every new solve

    if result["status"] == "error":
        st.error(f"**Hata:** {result['error']}")
    else:
        if result.get("warning"):
            st.warning(result["warning"])
        st.success("✓ Optimizasyon tamamlandı!")


# ===========================================================================
# PAGE: DIŞA AKTARMA
# ===========================================================================
def page_export():
    section_header("💾", "Dışa Aktarma", "Sonuçları Excel veya CSV olarak indirin")

    result = st.session_state.solver_result
    if result is None:
        st.markdown(
            '<div class="warn-banner">⚠️ Henüz bir çözüm çalıştırılmadı. Önce optimizasyonu başlatın.</div>',
            unsafe_allow_html=True,
        )
        if st.button("⚙️  Çözüm Ayarlarına Git"):
            nav_to("settings"); st.rerun()
        return

    if result.get("status") == "error":
        st.markdown(
            '<div class="err-banner">❌ Çözüm hatası oluştu. Dışa aktarılacak veri yok.</div>',
            unsafe_allow_html=True,
        )
        return

    # Block export for infeasible Model 1 / Heuristic 1
    if (result.get("status") == "infeasible"
            and st.session_state.method in ("Model 1", "Heuristic Model 1")):
        st.warning(
            "Dışa aktarılacak geçerli bir çözüm bulunmuyor. "
            "Tip bazında önerilen değerleri kontrol edin veya Model 2 / Heuristic Model 2 kullanın."
        )
        if st.button("⚙️  Çözüm Ayarlarına Dön", key="export_back_settings"):
            st.session_state.optimization_step = 3; nav_to("optimization"); st.rerun()
        return

    roll_data = result.get("roll_data", [])
    if not roll_data:
        st.info("Kullanılan bobin verisi bulunamadı.")
        return

    data_dict = st.session_state.data_dict or {}
    TD = data_dict.get("TD", {})
    LD = data_dict.get("LD", {})
    orders_df = st.session_state.orders_df

    display_df = build_roll_display_df(roll_data, TD, LD)

    # Map numeric type IDs → paper type labels in "Jumbo Bobin Tipi" column
    _type_map_state = st.session_state.get("type_map") or {}
    _inv_type_map = {v: k for k, v in _type_map_state.items()}
    if not display_df.empty and "Jumbo Bobin Tipi" in display_df.columns:
        display_df["Jumbo Bobin Tipi"] = display_df["Jumbo Bobin Tipi"].apply(
            lambda x: _inv_type_map.get(x, x) if x != "—" else "—"
        )
    if not display_df.empty:
        display_df["Kullanılan Uzunluk"] = display_df["Kullanılan Uzunluk"].apply(fmt_cm)
        display_df["Kağıt Kaybı"]        = display_df["Kağıt Kaybı"].apply(fmt_cm)

    # ── Summary cards ──────────────────────────────────────────────────────
    st.markdown("#### 📊 Çözüm Özeti")
    c1, c2, c3 = st.columns(3)
    with c1:
        metric_card(fmt_n(result["total_paper_waste"]), "Kağıt Kaybı", sub="cm")
    with c2:
        metric_card(str(result["total_time_waste"]), "Zaman Kaybı")
    with c3:
        metric_card(str(result["used_reel_count"]), "Kullanılan Bobin")

    # ── Download options ────────────────────────────────────────────────────
    st.markdown("#### 📥 İndirme Seçenekleri")

    col_xl, col_csv, col_ord = st.columns(3)

    # ── Excel export ────────────────────────────────────────────────────────
    with col_xl:
        st.markdown("##### 📗 Excel Dosyası")
        st.caption("Özet, bobin detayları ve siparişler — formatlanmış .xlsx")
        try:
            import io as _io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            buf_xl = _io.BytesIO()
            wb = Workbook()

            red_fill  = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
            gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            bold_font = Font(bold=True, size=10)

            # Durum değerini Türkçeleştir
            _durum_tr = {
                "optimal":               "Optimal",
                "feasible":              "Uygun Çözüm",
                "infeasible":            "Çözüm Bulunamadı",
                "time_limit":            "Zaman Limitine Ulaşıldı",
                "time_limit_no_solution":"Zaman Limiti (Çözüm Yok)",
                "error":                 "Hata",
            }
            _durum_val = _durum_tr.get(result["status"], result["status"])

            # Sheet 1: Summary
            ws1 = wb.active
            ws1.title = "Özet"
            rows = [
                ("Parametre", "Değer"),
                ("Yöntem", st.session_state.method),
                ("Lambda (λ)", f"{st.session_state.lambda_val:.2f}"),
                ("Jumbo Bobin Uzunluğu (L)", fmt_cm(st.session_state.L)),
                ("Bıçak Sayısı (B)", str(st.session_state.B)),
                ("Durum", _durum_val),
                ("Toplam Kağıt Kaybı", fmt_cm(result["total_paper_waste"])),
                ("Toplam Zaman Kaybı", str(result["total_time_waste"])),
                ("Kullanılan Bobin Sayısı", str(result["used_reel_count"])),
                ("Çözüm Süresi", f"{fmt_n(result['runtime'])} sn"),
            ]
            for i, (k, v) in enumerate(rows, start=1):
                c1c = ws1.cell(row=i, column=1, value=k)
                c2c = ws1.cell(row=i, column=2, value=v)
                if i == 1:
                    for c in [c1c, c2c]:
                        c.fill = red_fill
                        c.font = hdr_font
                        c.alignment = Alignment(horizontal="center")
                else:
                    c1c.font = bold_font
                    if i % 2 == 0:
                        c1c.fill = gray_fill
                        c2c.fill = gray_fill
            ws1.column_dimensions["A"].width = 30
            ws1.column_dimensions["B"].width = 26

            # Sheet 2: Reel Detail
            if not display_df.empty:
                ws2 = wb.create_sheet("Bobin Detayları")
                headers = list(display_df.columns)
                for ci, h in enumerate(headers, start=1):
                    cell = ws2.cell(row=1, column=ci, value=h)
                    cell.fill = red_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center")
                for ri, row in display_df.iterrows():
                    for ci, val in enumerate(row, start=1):
                        c = ws2.cell(row=ri + 2, column=ci, value=val)
                        if (ri + 1) % 2 == 0:
                            c.fill = gray_fill
                for ci in range(1, len(headers) + 1):
                    ws2.column_dimensions[get_column_letter(ci)].width = 18
                ws2.column_dimensions[get_column_letter(len(headers))].width = 42

            # Sheet 3: Orders
            if orders_df is not None:
                ws3 = wb.create_sheet("Siparişler")
                ord_headers = ["Kağıt Tipi", "Uzunluk (cm)", "Miktar (adet)"]
                for ci, h in enumerate(ord_headers, start=1):
                    cell = ws3.cell(row=1, column=ci, value=h)
                    cell.fill = red_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center")
                for ri, row in orders_df.iterrows():
                    ws3.cell(row=ri + 2, column=1, value=str(row["paper_type"]))
                    ws3.cell(row=ri + 2, column=2,
                             value=float(row["length"]) if not pd.isna(row["length"]) else "")
                    ws3.cell(row=ri + 2, column=3,
                             value=int(row["quantity"]) if not pd.isna(row["quantity"]) else "")
                    if (ri + 1) % 2 == 0:
                        for ci in [1, 2, 3]:
                            ws3.cell(row=ri + 2, column=ci).fill = gray_fill
                for ci, w in [(1, 16), (2, 16), (3, 16)]:
                    ws3.column_dimensions[get_column_letter(ci)].width = w

            wb.save(buf_xl)
            buf_xl.seek(0)
            st.download_button(
                label="📗 Excel Olarak İndir (.xlsx)",
                data=buf_xl,
                file_name="alkim_kesim_sonuclari.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        except Exception as e:
            st.error(f"Excel oluşturulamadı: {e}")

    # ── CSV: Reel detail ────────────────────────────────────────────────────
    with col_csv:
        st.markdown("##### 📄 Bobin Detayları (CSV)")
        st.caption("Bobin bazlı karar tablosu — .csv")
        if not display_df.empty:
            csv_data = display_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="📄 CSV Olarak İndir",
                data=csv_data,
                file_name="alkim_bobin_detaylari.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ── CSV: Orders ─────────────────────────────────────────────────────────
    with col_ord:
        st.markdown("##### 📋 Sipariş Listesi (CSV)")
        st.caption("Kullanılan sipariş verisi — .csv")
        if orders_df is not None:
            orders_export_df = orders_df.rename(columns={
                "paper_type": "Kağıt Tipi",
                "length":     "Uzunluk (cm)",
                "quantity":   "Miktar (adet)",
            })
            orders_csv = orders_export_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="📋 Siparişleri İndir (.csv)",
                data=orders_csv,
                file_name="alkim_siparisler.csv",
                mime="text/csv",
                use_container_width=True,
            )


    # ── Preview table ───────────────────────────────────────────────────────
    if not display_df.empty:
        st.markdown("#### 📋 Dışa Aktarılacak Bobin Tablosu (Önizleme)")
        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True,
            height=min(500, 80 + len(display_df) * 38),
        )

    # ── Raw solver output ───────────────────────────────────────────────────
    raw_out = result.get("raw_output", "")
    if raw_out and raw_out.strip():
        with st.expander("🔍 Çözücü Log Çıktısı"):
            st.code(raw_out, language="")

# ===========================================================================
# PAGE: SONUÇLAR
# ===========================================================================
def page_results():
    section_header("📊", "Sonuçlar", "Optimizasyon özeti ve bobin bazlı karar tablosu")

    result = st.session_state.solver_result
    if result is None:
        st.markdown(
            '<div class="warn-banner">⚠️ Henüz bir çözüm çalıştırılmadı. Çözüm Ayarları sayfasına gidin.</div>',
            unsafe_allow_html=True,
        )
        if st.button("⚙️  Kesim Optimizasyonuna Git"):
            st.session_state.optimization_step = 3; nav_to("optimization"); st.rerun()
        return

    # Error state
    if result["status"] == "error":
        st.markdown(
            f'<div class="err-banner">❌ <b>Çözüm Hatası</b><br><pre style="margin-top:8px;font-size:.8rem;white-space:pre-wrap">{result["error"]}</pre></div>',
            unsafe_allow_html=True,
        )
        return

    # Warning banner (time limit)
    if result.get("warning"):
        st.markdown(
            f'<div class="warn-banner">⚠️ {result["warning"]}</div>', unsafe_allow_html=True
        )

    # ── Reel count mismatch warning ────────────────────────────────────────
    _data_dict_chk      = st.session_state.data_dict or {}
    _user_reel_count    = len(_data_dict_chk.get("S", []))
    _used_reel_count    = result.get("used_reel_count") or 0

    if _used_reel_count > _user_reel_count > 0:
        if "reel_warn_dismissed" not in st.session_state:
            st.session_state["reel_warn_dismissed"] = False

        if not st.session_state["reel_warn_dismissed"]:
            _wcol, _bcol = st.columns([11, 1])
            with _wcol:
                st.warning(
                    f"**Girilen bobin sayısı ile çözüm mümkün değil**\n\n"
                    f"Girilen bobin sayısı: **{_user_reel_count}**  \n"
                    f"Gerekli minimum bobin sayısı: **{_used_reel_count}**\n\n"
                    f"Model daha yüksek bobin sayısı ile çözülmüştür."
                )
            with _bcol:
                st.markdown("<br><br>", unsafe_allow_html=True)
                if st.button("✕ Kapat", key="dismiss_reel_warn",
                             help="Uyarıyı kapat", use_container_width=True):
                    st.session_state["reel_warn_dismissed"] = True
                    st.rerun()

    # ── Model 1 / Heuristic 1 — infeasible: show error only, stop rendering ──
    if (result.get("status") == "infeasible"
            and st.session_state.method in ("Model 1", "Heuristic Model 1")):
        st.markdown(
            '<div style="background:#FEF2F2;border:1.5px solid #FCA5A5;border-radius:12px;'
            'padding:1.5rem 1.8rem;margin:1rem 0">'
            '<div style="font-size:1.15rem;font-weight:700;color:#DC2626;margin-bottom:0.6rem">'
            '❌ Çözüm bulunamadı</div>'
            '<div style="color:#374151;font-size:0.95rem;line-height:1.7">'
            'Girilen bobin tipi dağılımı yetersiz görünüyor.<br>'
            'Tip bazında önerilen değerleri kontrol edin veya '
            '<b>Model 2 / Heuristic Model 2</b> kullanın.'
            '</div></div>',
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("⚙️  Çözüm Ayarlarına Dön", type="primary",
                     key="infeasible_back_to_settings"):
            st.session_state.optimization_step = 3; nav_to("optimization")
            st.rerun()
        return

    # ── Status logic: if roll_data has used reels, always show success ─────
    _roll_data_all = result.get("roll_data", [])
    _used_count = sum(1 for r in _roll_data_all if r.get("used", False))
    display_status = result["status"]
    if display_status == "infeasible" and _used_count > 0:
        display_status = "feasible"

    # ── Summary cards ──────────────────────────────────────────────────────
    st.markdown(
        f'<div style="margin-bottom:1rem">Durum: {status_badge(display_status)} &nbsp;&nbsp; '
        f'Yöntem: <b>{st.session_state.method}</b> &nbsp;&nbsp; '
        f'Süre: <b>{fmt_n(result["runtime"])} sn</b></div>',
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        metric_card(fmt_n(result["total_paper_waste"]), "Toplam Kağıt Kaybı", sub="cm")
    with c2:
        metric_card(str(result["total_time_waste"]), "Toplam Zaman Kaybı", sub="bobin")
    with c3:
        metric_card(str(result["used_reel_count"]), "Kullanılan Bobin Sayısı")
    with c4:
        metric_card(fmt_n(result["runtime"]), "Çözüm Süresi", sub="saniye")

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    # ── Öncelik & Üretim Sırası Panelleri ─────────────────────────────────
    # State init
    st.session_state.setdefault("_pp_open", False)
    st.session_state.setdefault("_pp_mode", None)

    _cur_lam  = float(st.session_state.lambda_val)
    _cur_mode = ("paper" if _cur_lam >= 1.0 else "time" if _cur_lam <= 0.0 else "balanced")
    _kp_pct   = int(round(_cur_lam * 100))
    _zp_pct   = 100 - _kp_pct

    _show_seq = st.session_state.get("strategy_mode") == "given_schedule"

    # İki paneli 2:2 kolon düzeninde yan yana yerleştir.
    # Sağ kolon: given_schedule ise Üretim Sırası, değilse boş kalır.
    _panel_l, _panel_r = st.columns(2)

    # ── SOL PANEL: Çözüm Öncelikleri ──────────────────────────────────────
    with _panel_l:
        with st.container(border=True):
            if not st.session_state["_pp_open"]:
                # ── KAPALI HAL ─────────────────────────────────────────
                st.markdown(
                    '<span style="font-size:.85rem;font-weight:700;color:#1E3A5F;">'
                    '🎯 Çözüm Öncelikleri</span>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f'<div style="font-size:.82rem;color:#374151;margin:.25rem 0 .4rem 0;">'
                    f'Kağıt Kaybı Ağırlığı: <b>%{_kp_pct}</b> &nbsp;·&nbsp; '
                    f'Zaman Kaybı Ağırlığı: <b>%{_zp_pct}</b></div>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    '<div style="font-size:.79rem;color:#6B7280;margin-bottom:.5rem;">'
                    'Farklı bir öncelik denemek ister misiniz?</div>',
                    unsafe_allow_html=True,
                )
                if st.button("Evet, farklı öncelik dene →",
                             key="_pp_open_btn", use_container_width=True):
                    st.session_state["_pp_open"] = True
                    st.session_state["_pp_mode"] = _cur_mode
                    st.rerun()

            else:
                # ── AÇIK HAL ──────────────────────────────────────────
                st.markdown(
                    '<span style="font-size:.85rem;font-weight:700;color:#1E3A5F;">'
                    '🎯 Çözüm Önceliği Seçin</span>',
                    unsafe_allow_html=True,
                )

                _sel = st.session_state["_pp_mode"] or _cur_mode

                # ── 3 seçenek ─────────────────────────────────────────
                _opts = [
                    ("paper",    "📄 Kağıt Kaybını Azalt",  "Kağıt israfını önceliklendirir.",          "_pp_opt_paper"),
                    ("time",     "⏱ Zaman Kaybını Azalt",   "Bıçak düzeni değişimlerini azaltır.",      "_pp_opt_time"),
                    ("balanced", "⚖️ Dengeli Çözüm",         "İki hedef arasında denge kurar.",          "_pp_opt_balanced"),
                ]
                _clicked_mode = None
                for _omode, _otitle, _odesc, _okey in _opts:
                    _active = (_sel == _omode)
                    _badge  = (
                        ' <span style="font-size:.68rem;background:#C8102E;color:white;'
                        'padding:1px 7px;border-radius:10px;vertical-align:middle;">'
                        'Aktif Seçim</span>' if _active else ""
                    )
                    st.markdown(
                        f'<div style="font-size:.8rem;font-weight:600;color:#1E3A5F;'
                        f'margin:.55rem 0 2px 0;">{_otitle}{_badge}</div>'
                        f'<div style="font-size:.75rem;color:#6B7280;margin-bottom:3px;">{_odesc}</div>',
                        unsafe_allow_html=True,
                    )
                    if st.button(
                        "✓ Seçili" if _active else "Seç",
                        key=_okey,
                        use_container_width=True,
                        type="primary" if _active else "secondary",
                    ):
                        _clicked_mode = _omode

                if _clicked_mode is not None:
                    st.session_state["_pp_mode"] = _clicked_mode
                    st.rerun()

                # ── Dengeli Çözüm slider ───────────────────────────────
                if _sel == "balanced":
                    st.markdown(
                        '<div style="font-size:.75rem;color:#6B7280;margin:.5rem 0 .2rem 0;">'
                        'Kaydırıcıyı değiştirerek çözümü kağıt kaybı veya zaman kaybı '
                        'tarafına yaklaştırabilirsiniz.</div>',
                        unsafe_allow_html=True,
                    )
                    _bal_default = max(0.01, min(0.99,
                        _cur_lam if _cur_mode == "balanced" else 0.5))
                    _bal_v = st.slider(
                        "Kağıt Kaybı Ağırlığı",
                        min_value=0.01, max_value=0.99,
                        value=st.session_state.get("_pp_bal_slider", _bal_default),
                        step=0.01,
                        key="_pp_bal_slider",
                    )
                    _kp2 = int(round(_bal_v * 100))
                    st.caption(f"Kağıt Kaybı %{_kp2}  —  Zaman Kaybı %{100 - _kp2}")

                # ── Alt butonlar ───────────────────────────────────────
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                _rb1, _rb2 = st.columns([3, 2])
                with _rb1:
                    if st.button("🔁 Tekrar Optimize Et",
                                 key="_pp_rerun", type="primary",
                                 use_container_width=True):
                        # Seçilen moda göre gerçek lambda belirle
                        if _sel == "paper":
                            _new_lam = 1.0
                        elif _sel == "time":
                            _new_lam = 0.0
                        else:
                            _new_lam = float(st.session_state.get("_pp_bal_slider", 0.5))
                        # Uygula ve tekrar çöz
                        st.session_state.lambda_val = _new_lam
                        st.session_state["_pp_open"] = False
                        st.session_state["_pp_mode"] = None
                        # Slider'ı sıfırla, bir sonraki açılışta yeni lambda'yı esas alsın
                        st.session_state.pop("_pp_bal_slider", None)
                        # Yeni çözüm gelince sıra-iyileştirme geçmişi geçersiz olur
                        st.session_state["_seq_opt_active"] = False
                        st.session_state.pop("_prev_gs_result", None)
                        st.session_state.pop("_prev_gs_method", None)
                        st.session_state.pop("_prev_gs_tj", None)
                        _run_solver()
                        st.rerun()
                with _rb2:
                    if st.button("Kapat", key="_pp_close",
                                 use_container_width=True):
                        st.session_state["_pp_open"] = False
                        st.session_state["_pp_mode"] = None
                        st.rerun()

    # ── SAĞ PANEL: Üretim Sırası (sadece given_schedule stratejisinde) ────
    with _panel_r:
        if _show_seq:
            st.session_state.setdefault("_seq_opt_active", False)
            with st.container(border=True):
                _seq_active = st.session_state["_seq_opt_active"]

                if not _seq_active:
                    # ── Normal hal: sıra iyileştirme teklifi ──────────
                    st.markdown(
                        '<span style="font-size:.85rem;font-weight:700;color:#1E3A5F;">'
                        '🔀 Üretim Sırası</span>',
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        '<div style="font-size:.79rem;color:#374151;margin:.25rem 0 .4rem 0;">'
                        'Bu çözümde bobinlerin üretim sırası kullanıcı tarafından verilmiştir. '
                        'Sistem yalnızca verilen sıraya göre kesim desenlerini optimize etmiştir.'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        '<div style="font-size:.79rem;color:#6B7280;margin-bottom:.4rem;">'
                        'Bobin sırasını da sistemin iyileştirmesini ister misiniz?</div>',
                        unsafe_allow_html=True,
                    )
                    st.caption(
                        "Sistem, mevcut çözüm yapısına uygun otomatik "
                        "sıralama yöntemini kullanacaktır."
                    )
                    if st.button("🔀 Sırayı Sistem İyileştirsin",
                                 key="_seq_opt_run", use_container_width=True,
                                 type="secondary"):
                        st.session_state["_prev_gs_result"] = st.session_state.solver_result
                        st.session_state["_prev_gs_method"] = st.session_state.method
                        st.session_state["_prev_gs_tj"]     = st.session_state.tj_table
                        _gs_cur_method = st.session_state.method
                        _gs_new_method = (
                            "Model 2" if _gs_cur_method == "Model 1"
                            else "Heuristic Model 2"
                        )
                        st.session_state.method             = _gs_new_method
                        st.session_state.tj_table           = None
                        st.session_state["_seq_opt_active"] = True
                        _run_solver()
                        st.rerun()

                else:
                    # ── Sıra iyileştirilmiş hal: bilgi + geri dön ─────
                    st.markdown(
                        '<span style="font-size:.85rem;font-weight:700;color:#1E3A5F;">'
                        '🔀 Üretim Sırası</span>',
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        '<div style="font-size:.79rem;background:#F0FDF4;'
                        'border:1px solid #86EFAC;border-radius:8px;'
                        'padding:6px 10px;margin:.25rem 0 .5rem 0;color:#166534;">'
                        '✓ Bu sonuçta bobin sırası sistem tarafından iyileştirilmiştir.'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                    if st.button("↩ Verilen Sıraya Geri Dön",
                                 key="_seq_rollback", use_container_width=True):
                        _prev_r = st.session_state.get("_prev_gs_result")
                        _prev_m = st.session_state.get("_prev_gs_method")
                        _prev_t = st.session_state.get("_prev_gs_tj")
                        if _prev_r is not None:
                            st.session_state.solver_result = _prev_r
                            st.session_state.method        = _prev_m
                            st.session_state.tj_table      = _prev_t
                        st.session_state["_seq_opt_active"] = False
                        st.rerun()

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # ── Roll-level table ───────────────────────────────────────────────────
    st.markdown("#### 📋 Bobin Bazlı Karar Tablosu")

    roll_data = result.get("roll_data", [])
    data_dict = st.session_state.data_dict or {}
    TD = data_dict.get("TD", {})
    LD = data_dict.get("LD", {})

    if roll_data and TD and LD:
        display_df = build_roll_display_df(roll_data, TD, LD)
        # Map numeric type IDs → paper type labels
        _type_map_s2 = st.session_state.get("type_map") or {}
        _inv_type_map_s2 = {v: k for k, v in _type_map_s2.items()}
        if not display_df.empty and "Jumbo Bobin Tipi" in display_df.columns:
            display_df["Jumbo Bobin Tipi"] = display_df["Jumbo Bobin Tipi"].apply(
                lambda x: _inv_type_map_s2.get(x, x) if x != "—" else "—"
            )
        if not display_df.empty:
            display_df["Kullanılan Uzunluk"] = display_df["Kullanılan Uzunluk"].apply(fmt_cm)
            display_df["Kağıt Kaybı"]        = display_df["Kağıt Kaybı"].apply(fmt_cm)
            st.dataframe(
                display_df,
                use_container_width=True,
                hide_index=True,
                height=min(600, 80 + len(display_df) * 38),
                column_config={
                    "Jumbo Bobin No":      st.column_config.NumberColumn(format="%d", width="small"),
                    "Jumbo Bobin Tipi":    st.column_config.TextColumn(width="small"),
                    "Kesim Deseni":       st.column_config.TextColumn(width="medium"),
                    "Kullanılan Uzunluk": st.column_config.TextColumn(width="small"),
                    "Bıçak Mesafeleri":   st.column_config.TextColumn(width="medium"),
                    "Bıçak Sayısı":       st.column_config.NumberColumn(format="%d", width="small"),
                    "Kağıt Kaybı":        st.column_config.TextColumn(width="small"),
                    "Zaman Kaybı":        st.column_config.NumberColumn(format="%d", width="small"),
                    "Atanan Siparişler":  st.column_config.TextColumn(width="large"),
                },
            )
        else:
            st.info("Kullanılan bobin bulunamadı.")
    else:
        st.info("Bobin verisi mevcut değil.")


    # ── Parameter summary ──────────────────────────────────────────────────
    with st.expander("🔧 Kullanılan Parametre Özeti"):
        pc1, pc2, pc3, pc4 = st.columns(4)
        with pc1:
            st.markdown(f"**Yöntem**<br>{st.session_state.method}", unsafe_allow_html=True)
        with pc2:
            st.markdown(f"**λ (Lambda)**<br>{st.session_state.lambda_val:.2f}", unsafe_allow_html=True)
        with pc3:
            st.markdown(f"**L (Bobin Boyu)**<br>{fmt_cm(st.session_state.L)}", unsafe_allow_html=True)
        with pc4:
            st.markdown(f"**B (Bıçak Sayısı)**<br>{st.session_state.B}", unsafe_allow_html=True)

    # ── Raw solver output ──────────────────────────────────────────────────
    raw_out = result.get("raw_output", "")
    if raw_out and raw_out.strip():
        with st.expander("🔍 Çözücü Log Çıktısı"):
            st.code(raw_out, language="")

    # ── Navigation ────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    col_nav1, col_nav2, _ = st.columns([1, 1, 2])
    with col_nav1:
        if st.button("📈  Görsel Analize Git →", type="primary",
                     use_container_width=True, key="results_to_analysis"):
            st.session_state.active_page = "analysis"
            st.rerun()
    with col_nav2:
        if st.button("💾  Dışa Aktarmaya Git →",
                     use_container_width=True, key="results_to_export"):
            st.session_state.active_page = "export"
            st.rerun()


# ===========================================================================
# PAGE: GÖRSEL ANALİZ
# ===========================================================================
def page_analysis():
    section_header("📈", "Görsel Analiz", "Sonuçların grafiksel değerlendirmesi")

    result = st.session_state.solver_result
    if result is None:
        st.markdown(
            '<div class="warn-banner">⚠️ Henüz bir çözüm çalıştırılmadı. Çözüm Ayarları sayfasına gidin.</div>',
            unsafe_allow_html=True,
        )
        if st.button("⚙️  Çözüm Ayarlarına Git"):
            nav_to("settings"); st.rerun()
        return

    if result.get("status") == "error":
        st.markdown(
            '<div class="err-banner">❌ Çözüm hatası oluştu. Sonuç görüntülenemiyor.</div>',
            unsafe_allow_html=True,
        )
        return

    # Block analysis for infeasible Model 1 / Heuristic 1
    if (result.get("status") == "infeasible"
            and st.session_state.method in ("Model 1", "Heuristic Model 1")):
        st.warning(
            "Görsel analiz için geçerli bir çözüm bulunmuyor. "
            "Tip bazında önerilen değerleri kontrol edin veya Model 2 / Heuristic Model 2 kullanın."
        )
        if st.button("⚙️  Çözüm Ayarlarına Dön", key="analysis_back_settings"):
            st.session_state.optimization_step = 3; nav_to("optimization"); st.rerun()
        return

    roll_data  = result.get("roll_data", [])
    used_rolls = [r for r in roll_data if r.get("used", False)]
    if not used_rolls:
        st.info("Kullanılan bobin verisi bulunamadı.")
        return

    data_dict = st.session_state.data_dict or {}
    LD    = data_dict.get("LD", {})
    TD    = data_dict.get("TD", {})
    L_val = float(data_dict.get("L", 348))

    # ── Alkım color theme ──────────────────────────────────────────────────
    CLR_USED  = "#1F2937"   # dark charcoal  — used length
    CLR_WASTE = "#C8102E"   # Alkım red      — paper waste
    CLR_TW    = "#6B7280"   # neutral gray   — time waste
    CLR_GRID  = "#F3F4F6"   # off-white      — plot background

    # ── PAPER TYPE COLOR MAPPING ───────────────────────────────────────────
    FIXED_PAPER_COLORS = [
        "#EFE4CF", "#E4D2B3", "#D8C09A", "#CDAE81", "#BD9768",
        "#AB815B", "#9B7F6A", "#887B70", "#756C65", "#655D58"
    ]

    # ── inv_type_map: int solver ID → paper label (built first, used everywhere below) ──
    _type_map_anal = st.session_state.get("type_map") or {}
    _inv_type_map_anal = {v: k for k, v in _type_map_anal.items()}  # {int_id: label_str}

    def _sort_type_key(t):
        """Always returns a (group, value) tuple so sorted() never compares float vs str."""
        s = str(t).strip()
        if not s or s.lower() == "nan":
            return (2, "")
        # Numeric string or int/float → sort numerically in group 0
        try:
            return (0, float(s))
        except (ValueError, TypeError):
            pass
        # Label string → map to numeric ID for consistent ordering in group 0
        if s in PAPER_TYPE_IDS:
            return (0, float(PAPER_TYPE_IDS[s]))
        # Unknown string → sort lexicographically in group 1
        return (1, s)

    orders_df = st.session_state.orders_df
    if orders_df is not None:
        # Normalize: convert any residual numeric IDs → labels, drop None/NaN/empty
        _raw_types = orders_df["paper_type"].astype(str).unique()
        _all_types_set = set()
        for _tv in _raw_types:
            _tv = str(_tv).strip()
            if not _tv or _tv.lower() == "nan":
                continue
            # If it's a numeric string, try to resolve to a label
            try:
                _tv = _inv_type_map_anal.get(int(float(_tv)), _tv)
            except (ValueError, TypeError):
                pass
            _all_types_set.add(_tv)
        _all_types = sorted(_all_types_set, key=_sort_type_key)
    else:
        _all_types = []

    # Get unique types used in the solver result to ensure we map them too
    all_demand_ids = set()
    for r in used_rolls:
        all_demand_ids.update(r.get("assignments", {}).keys())
        if r.get("last_demand") is not None:
            all_demand_ids.add(r["last_demand"])
    all_demand_ids = sorted(all_demand_ids)
    unique_types_used = sorted(set(TD.get(d, 0) for d in all_demand_ids), key=_sort_type_key)

    # Merge solver types into _all_types using labels (not raw int strings)
    _all_types_set2 = set(_all_types)
    for _t in unique_types_used:
        _lbl = _inv_type_map_anal.get(_t, None)
        if _lbl and str(_lbl).strip() and str(_lbl).lower() != "nan":
            _all_types_set2.add(_lbl)
        elif _lbl is None:
            # No label mapping — fall back to str(t) but only if not already present as a label
            _fallback = str(_t)
            if not any(_sort_type_key(x) == _sort_type_key(_fallback) for x in _all_types_set2):
                _all_types_set2.add(_fallback)

    _all_types = sorted(_all_types_set2, key=_sort_type_key)

    PAPER_TYPE_COLORS = {}
    import colorsys
    for i, t in enumerate(_all_types):
        if i < len(FIXED_PAPER_COLORS):
            PAPER_TYPE_COLORS[str(t)] = FIXED_PAPER_COLORS[i]
        else:
            # Generate distinct colors avoiding red and black
            hue = 0.15 + (0.7 * ((i - len(FIXED_PAPER_COLORS)) / max(1, len(_all_types) - len(FIXED_PAPER_COLORS))))
            sat = 0.4 + (i % 3) * 0.15
            light = 0.6 + (i % 2) * 0.15
            r, g, b = colorsys.hls_to_rgb(hue, light, sat)
            PAPER_TYPE_COLORS[str(t)] = f"#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}"

    # Map integer type IDs → colors using the label-keyed PAPER_TYPE_COLORS
    type_color_map = {}
    for t in unique_types_used:
        lbl = _inv_type_map_anal.get(t, str(t))
        type_color_map[t] = PAPER_TYPE_COLORS.get(lbl, PAPER_TYPE_COLORS.get(str(t), "#6b7280"))

    reel_labels = [f"Bobin {r['reel_id']}" for r in used_rolls]

    tab_genel, tab_kesim = st.tabs(["📉 Genel Analiz", "✂️ Kesim Diyagramı"])

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — GENEL ANALİZ
    # ══════════════════════════════════════════════════════════════════════════
    with tab_genel:

        # ══════════════════════════════════════════════════════════════════
        # BÖLÜM 1 — Girdilerle İlgili İstatistikler
        # ══════════════════════════════════════════════════════════════════
        st.markdown(
            '<div style="border-top:2px solid #1E3A5F;margin-bottom:1rem;padding-top:.9rem;">'
            '<span style="font-size:1.35rem;font-weight:800;color:#1E3A5F;'
            'letter-spacing:.01em;">📊&nbsp; Girdilerle İlgili İstatistikler</span>'
            '<div style="font-size:.85rem;font-weight:400;color:#6B7280;margin-top:4px;">'
            'Bu bölümde sisteme girilen sipariş verilerine ait dağılımlar gösterilmektedir.'
            '</div></div>',
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        # ── Talep Dağılımı ────────────────────────────────────────────────
        try:
            orders_df = st.session_state.orders_df
            type_qty  = {}
            type_len  = {}
            if orders_df is not None:
                for _, row in orders_df.iterrows():
                    pt  = str(row["paper_type"])
                    qty = int(row["quantity"]) if not pd.isna(row["quantity"]) else 0
                    ln  = float(row["length"])  if not pd.isna(row["length"])  else 0.0
                    type_qty[pt] = type_qty.get(pt, 0) + qty
                    type_len[pt] = type_len.get(pt, 0) + qty * ln

            if type_qty:
                st.markdown("#### 🎯 Talep Dağılımı")
                types_list = sorted(type_qty.keys())
                qty_vals   = [type_qty[t] for t in types_list]
                len_vals   = [type_len[t] for t in types_list]
                pie_colors = [PAPER_TYPE_COLORS.get(str(t), "#E5E7EB") for t in types_list]

                n_types = len(types_list)
                # Taller charts when many types to avoid crowding
                _chart_h = max(400, 300 + max(0, n_types - 4) * 20)

                # ── Donut chart short-label mapping (longer than PAPER_TYPE_SHORT
                #    which is kept compact for cutting diagram segments)
                _PIE_LABEL_SHORT = {
                    "1.HAMUR OFSET":                 "OFSET",
                    "1.HAMUR IVORY(D)":              "IVORY(D)",
                    "1.HAMUR BARDAK":                "BARDAK",
                    "1.HAMUR AMBALAJ KAGIDI":        "AMBALAJ KAGIDI",
                    "FSC MIX CREDIT KRAFT ÇANTA OB": "KRAFT ÇANTA OB",
                    "1.HAMUR BRISTOL":               "BRISTOL",
                    "1.HAMUR KRAFT CANTA":           "KRAFT CANTA",
                    "1.HAMUR CEVAP KAGIDI":          "CEVAP KAGIDI",
                }
                def _pie_short(lbl):
                    s = str(lbl)
                    return _PIE_LABEL_SHORT.get(s, PAPER_TYPE_SHORT.get(s, s))

                _full_labels  = [str(t) for t in types_list]
                _short_labels = [_pie_short(t) for t in types_list]

                # Dynamic height: more space for more types (external labels need room)
                _chart_h_pie = max(520, 420 + max(0, n_types - 3) * 28)
                # Legend rows (3 items/row), bottom margin to keep legend below chart
                _pie_leg_rows = max(1, math.ceil(n_types / 3))
                _pie_leg_b    = max(80, _pie_leg_rows * 28 + 30)

                col_pie, col_bar = st.columns(2)
                with col_pie:
                    st.caption("Sipariş adedi dağılımı (kağıt tipine göre)")
                    fig_pie = go.Figure(go.Pie(
                        # labels = short names (shown in legend + used in texttemplate)
                        labels=_short_labels,
                        values=qty_vals,
                        # customdata carries full names for hover tooltip
                        customdata=_full_labels,
                        marker=dict(colors=pie_colors, line=dict(color="white", width=1.5)),
                        hole=0.42,
                        # External label: "19.3%\nOFSET" — percent + short name, nothing inside
                        text=_short_labels,
                        texttemplate="%{percent:.1%}<br><b>%{text}</b>",
                        textposition="outside",
                        # Color each label to match its slice → acts as the color indicator
                        textfont=dict(color=pie_colors, size=10,
                                      family="Inter, Arial, sans-serif"),
                        automargin=True,
                        # Hover: full name + quantity + percentage
                        hovertemplate=(
                            "<b>%{customdata}</b><br>"
                            "Adet: %{value}<br>"
                            "Oran: %{percent:.1%}"
                            "<extra></extra>"
                        ),
                    ))
                    fig_pie.update_layout(
                        height=_chart_h_pie,
                        paper_bgcolor="white",
                        margin=dict(l=20, r=20, t=20, b=_pie_leg_b),
                        showlegend=True,
                        legend=dict(
                            orientation="h",
                            yanchor="top",
                            y=-0.03,
                            xanchor="center",
                            x=0.5,
                            font=dict(size=10, family="Inter, Arial, sans-serif"),
                            # ~3 entries per row via entrywidth fraction
                            entrywidth=0.32,
                            entrywidthmode="fraction",
                            itemsizing="constant",
                            tracegroupgap=6,
                        ),
                    )
                    st.plotly_chart(fig_pie, use_container_width=True)

                with col_bar:
                    st.caption("Toplam talep uzunluğu (cm, kağıt tipine göre)")
                    # Rotate x-axis labels when many types so names don't overlap
                    _tick_angle = 0 if n_types <= 4 else (-35 if n_types <= 7 else -55)
                    _b_margin   = max(60, 30 + n_types * 8) if n_types > 4 else 40
                    fig_bar2 = go.Figure(go.Bar(
                        x=[str(t) for t in types_list],
                        y=len_vals,
                        marker_color=pie_colors,
                        marker_line_width=0,
                        text=[f"{v:,.0f}" for v in len_vals],
                        textposition="outside",
                        textfont=dict(size=10),
                        hovertemplate="<b>%{x}</b><br>Uzunluk: %{y:g} cm<extra></extra>",
                    ))
                    fig_bar2.update_layout(
                        height=_chart_h_pie, bargap=0.4,
                        plot_bgcolor=CLR_GRID, paper_bgcolor="white",
                        margin=dict(l=10, r=20, t=50, b=_b_margin),
                        yaxis_title="Toplam Uzunluk (cm)",
                        yaxis=dict(gridcolor="white"),
                        xaxis=dict(
                            gridcolor="white",
                            tickangle=_tick_angle,
                            automargin=True,
                        ),
                        uniformtext_minsize=9,
                        uniformtext_mode="hide",
                    )
                    st.plotly_chart(fig_bar2, use_container_width=True)
        except Exception as e:
            st.error(f"Talep dağılımı grafiği oluşturulamadı: {e}")

        st.markdown("<br><br>", unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════
        # BÖLÜM 2 — Çözümle İlgili İstatistikler
        # ══════════════════════════════════════════════════════════════════
        st.markdown(
            '<div style="border-top:2px solid #C8102E;margin-bottom:1rem;padding-top:.9rem;">'
            '<span style="font-size:1.35rem;font-weight:800;color:#C8102E;'
            'letter-spacing:.01em;">✂️&nbsp; Çözümle İlgili İstatistikler</span>'
            '<div style="font-size:.85rem;font-weight:400;color:#6B7280;margin-top:4px;">'
            'Bu bölümde optimizasyon sonucunda oluşan kesim planına ait istatistikler gösterilmektedir.'
            '</div></div>',
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        # ── 1. Bobin Bazlı Kağıt Kullanımı ve Kaybı ──────────────────────
        try:
            pw_vals       = [r.get("paper_waste", 0.0)  for r in used_rolls]
            used_len_vals = [r.get("used_length", 0.0)  for r in used_rolls]

            st.markdown("#### 📏 Bobin Bazlı Kağıt Kullanımı ve Kaybı")
            st.caption("Her bobin için kullanılan uzunluk ve kağıt kaybı.")

            fig_pw = go.Figure()
            fig_pw.add_trace(go.Bar(
                name="Kullanılan Uzunluk",
                y=reel_labels, x=used_len_vals,
                orientation="h", marker_color="#336F4D", marker_line_width=0,
                hovertemplate="<b>%{y}</b><br>Kullanılan: %{x:g} cm<extra></extra>",
            ))
            fig_pw.add_trace(go.Bar(
                name="Kağıt Kaybı",
                y=reel_labels, x=pw_vals,
                orientation="h", marker_color=CLR_WASTE, marker_line_width=0,
                hovertemplate="<b>%{y}</b><br>Kayıp: %{x:g} cm<extra></extra>",
            ))
            fig_pw.update_layout(
                barmode="stack",
                height=max(320, len(used_rolls) * 34 + 100),
                plot_bgcolor=CLR_GRID, paper_bgcolor="white",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
                margin=dict(l=10, r=60, t=45, b=15),
                xaxis_title="Uzunluk (cm)",
                xaxis=dict(range=[0, L_val * 1.06], gridcolor="white"),
                yaxis=dict(autorange="reversed", gridcolor="white"),
            )
            fig_pw.add_vline(
                x=L_val, line_dash="dash", line_color="#C8102E", line_width=1.5,
                annotation_text=f"L = {fmt_n(L_val)} cm",
                annotation_font_color="#C8102E",
                annotation_position="top right",
            )
            st.plotly_chart(fig_pw, use_container_width=True)
        except Exception as e:
            st.error(f"Kağıt kaybı grafiği oluşturulamadı: {e}")

        st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)

        # ── 2. Kesim Deseni Tekrar Sayıları ──────────────────────────────
        try:
            pattern_freq = {}
            for r in used_rolls:
                asgn  = r.get("assignments", {})
                parts = []
                for d, cnt in sorted(asgn.items()):
                    parts.extend([f"{fmt_n(LD.get(d,'?'))}cm"] * cnt)
                if r.get("last_demand") is not None:
                    parts.append(f"{fmt_n(LD.get(r['last_demand'],'?'))}cm[son]")
                pat_key = " + ".join(parts) if parts else "(boş)"
                pattern_freq[pat_key] = pattern_freq.get(pat_key, 0) + 1

            if pattern_freq:
                st.markdown("#### 🔁 Kesim Deseni Tekrar Sayıları")
                st.caption("Aynı kesim desenine sahip bobin sayısı")

                sorted_pats = sorted(pattern_freq.items(), key=lambda x: -x[1])
                pat_labels  = [p for p, _ in sorted_pats]
                pat_counts  = [c for _, c in sorted_pats]

                fig_pat = go.Figure(go.Bar(
                    y=pat_labels, x=pat_counts,
                    orientation="h", marker_color=CLR_USED, marker_line_width=0,
                    text=pat_counts, textposition="outside",
                    hovertemplate="<b>%{y}</b><br>Tekrar: %{x}<extra></extra>",
                ))
                fig_pat.update_layout(
                    height=max(260, len(sorted_pats) * 40 + 90),
                    plot_bgcolor=CLR_GRID, paper_bgcolor="white",
                    margin=dict(l=10, r=60, t=15, b=25),
                    xaxis_title="Bobin Sayısı",
                    xaxis=dict(gridcolor="white"),
                    yaxis=dict(autorange="reversed", gridcolor="white"),
                )
                st.plotly_chart(fig_pat, use_container_width=True)
        except Exception as e:
            st.error(f"Desen tekrarı grafiği oluşturulamadı: {e}")

        st.markdown("<hr class='dss-divider'>", unsafe_allow_html=True)

        # ── 3. Bobin Bazlı Zaman Kaybı ───────────────────────────────────
        try:
            tw_vals = [r.get("time_waste", 0) for r in used_rolls]
            if any(v > 0 for v in tw_vals):
                st.markdown("#### ⏱ Bobin Bazlı Zaman Kaybı")
                st.caption("Bobinler arası bıçak düzeni değişimlerini gösterir.")

                change_count = sum(1 for v in tw_vals if v > 0)
                no_change    = len(tw_vals) - change_count
                n_reels      = len(tw_vals)

                reel_nums  = [r["reel_id"] for r in used_rolls]
                bar_clrs   = ["#C8102E" if v > 0 else "#D1D5DB" for v in tw_vals]
                hover_txts = []
                for rnum, val in zip(reel_nums, tw_vals):
                    if val > 0:
                        hover_txts.append(
                            f"<b>Bobin {rnum}</b><br>"
                            "Bıçak düzeni değişti<br>"
                            "→ Zaman kaybı oluştu"
                        )
                    else:
                        hover_txts.append(
                            f"<b>Bobin {rnum}</b><br>"
                            "Önceki bobin ile aynı bıçak düzeni<br>"
                            "→ Zaman kaybı yok"
                        )

                fig_tw = go.Figure()
                fig_tw.add_trace(go.Bar(
                    x=reel_nums,
                    y=[1] * n_reels,
                    marker_color=bar_clrs,
                    marker_line_width=0,
                    width=0.55,
                    hovertext=hover_txts,
                    hoverinfo="text",
                    showlegend=False,
                ))
                fig_tw.add_trace(go.Bar(
                    x=[None], y=[None],
                    marker_color="#C8102E",
                    name="Zaman kaybı var",
                    showlegend=True,
                ))
                fig_tw.add_trace(go.Bar(
                    x=[None], y=[None],
                    marker_color="#D1D5DB",
                    name="Zaman kaybı yok",
                    showlegend=True,
                ))
                fig_tw.update_layout(
                    height=130,
                    bargap=0.25,
                    plot_bgcolor="#F9FAFB",
                    paper_bgcolor="white",
                    margin=dict(l=10, r=10, t=10, b=45),
                    xaxis=dict(
                        title="Bobin No",
                        tickmode="linear",
                        tick0=reel_nums[0] if reel_nums else 1,
                        dtick=max(1, n_reels // 20),
                        gridcolor="white",
                        tickfont=dict(size=10),
                    ),
                    yaxis=dict(
                        showticklabels=False,
                        showgrid=False,
                        zeroline=False,
                        range=[0, 1.4],
                    ),
                    legend=dict(
                        orientation="h",
                        yanchor="bottom", y=1.04,
                        xanchor="right", x=1,
                        font=dict(size=10),
                        traceorder="normal",
                    ),
                    showlegend=True,
                )
                st.plotly_chart(fig_tw, use_container_width=True)

                st.markdown(
                    f'<div style="font-size:12px;color:#6B7280;margin-top:4px;'
                    f'padding:8px 12px;background:#F9FAFB;border-radius:8px;'
                    f'border-left:3px solid #C8102E">'
                    f'<b>Toplam zaman kaybı:</b> {change_count} bobin &nbsp;|&nbsp; '
                    f'<b>Zaman kaybı olmayan bobin sayısı:</b> {no_change} bobin'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        except Exception as e:
            st.error(f"Zaman kaybı grafiği oluşturulamadı: {e}")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — KESİM DİYAGRAMI  (full redesign: navy palette + per-segment labels)
    # ══════════════════════════════════════════════════════════════════════════
    with tab_kesim:
        try:
            st.markdown("#### ✂️ Jumbo Bobin Kesim Diyagramı")
            st.caption(
                "Her satır bir bobini temsil eder. "
                "Segmentler: **uzunluk | kağıt tipi**. "
                "Kırmızı blok = kağıt kaybı."
            )

            # ── Color helpers ──────────────────────────────────────────────
            WASTE_CLR = "#dc2626"  # Red — paper waste ONLY, never for types

            def _hex_to_rgb(h):
                h = h.lstrip("#")
                return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))

            def _perceived_brightness(hex_color):
                r, g, b = _hex_to_rgb(hex_color)
                return (r * 299 + g * 587 + b * 114) / 1000

            def _text_color_for_bg(hex_color):
                """White text on dark segments, dark text on light segments."""
                return "white" if _perceived_brightness(hex_color) < 140 else "#1F2937"

            cut_type_colors = type_color_map

            # ── Reel labels with Used / Waste summary ──────────────────────
            kesim_labels = [f"Bobin {r['reel_id']}" for r in used_rolls]

            # ── Build ordered segments per reel ────────────────────────────
            def _get_segments(r):
                segs = []
                for d in sorted(r.get("assignments", {}).keys()):
                    cnt = r["assignments"][d]
                    for _ in range(cnt):
                        ln = LD.get(d, 0)
                        tp = TD.get(d, 0)
                        segs.append((ln, tp, d, False))
                if r.get("last_demand") is not None:
                    ld_id = r["last_demand"]
                    segs.append((LD.get(ld_id, 0), TD.get(ld_id, 0), ld_id, True))
                return segs

            all_segs  = [_get_segments(r) for r in used_rolls]
            max_slots = max((len(s) for s in all_segs), default=0)
            n_reels   = len(used_rolls)

            # Minimum visual width for waste so tiny non-zero waste stays visible
            MIN_WASTE_VIS = L_val * 0.02

            fig_cut = go.Figure()

            # ── One trace per slot position ────────────────────────────────
            # Each slot = one cut position across all reels.  Per-reel color,
            # label and hover are passed as lists within the single trace.
            for slot in range(max_slots):
                x_vals    = []
                clr_list  = []
                txt_list  = []
                hov_list  = []
                txt_clrs  = []

                for ri in range(n_reels):
                    segs = all_segs[ri]
                    if slot < len(segs):
                        seg_len, seg_tp, seg_d, is_last = segs[slot]
                        x_vals.append(seg_len)
                        bg = cut_type_colors.get(seg_tp, "#6b7280")
                        clr_list.append(bg)
                        txt_clrs.append(_text_color_for_bg(bg))
                        # Label: "170 | OFSET" — shown only when segment is wide enough
                        _seg_lbl_full = _inv_type_map_anal.get(seg_tp, str(seg_tp))
                        _seg_short = PAPER_TYPE_SHORT.get(_seg_lbl_full, _seg_lbl_full[:8])
                        if seg_len >= L_val * 0.05:
                            lbl_len = int(seg_len) if seg_len == int(seg_len) else round(seg_len, 1)
                            txt_list.append(f"{lbl_len} | {_seg_short}")
                        else:
                            txt_list.append("")
                        last_tag = " (Son Dilim)" if is_last else ""
                        hov_list.append(
                            f"<b>Bobin {used_rolls[ri]['reel_id']}</b><br>"
                            f"Uzunluk: {fmt_n(seg_len)} cm<br>"
                            f"Tip: {_seg_lbl_full}{last_tag}<br>"
                            f"Kağıt kaybı: Hayır"
                        )
                    else:
                        x_vals.append(0)
                        clr_list.append("rgba(0,0,0,0)")
                        txt_list.append("")
                        hov_list.append("")
                        txt_clrs.append("white")

                fig_cut.add_trace(go.Bar(
                    y=kesim_labels,
                    x=x_vals,
                    orientation="h",
                    marker_color=clr_list,
                    marker_line_color="white",
                    marker_line_width=1.5,
                    text=txt_list,
                    textposition="inside",
                    textfont=dict(size=11, color=txt_clrs, family="Inter, sans-serif"),
                    textangle=0,
                    insidetextanchor="middle",
                    hovertext=hov_list,
                    hoverinfo="text",
                    showlegend=False,
                ))

            # ── Waste trace ────────────────────────────────────────────────
            waste_actual  = [max(0.0, L_val - float(r.get("used_length", 0))) for r in used_rolls]
            # Enforce minimum visual width so tiny waste stays red & visible
            waste_display = [max(w, MIN_WASTE_VIS) if w > 0 else 0.0 for w in waste_actual]
            waste_txt = []
            for wa in waste_actual:
                if wa >= L_val * 0.05:
                    wl = int(wa) if wa == int(wa) else round(wa, 1)
                    waste_txt.append(f"Kayıp {wl}")
                else:
                    waste_txt.append("")

            fig_cut.add_trace(go.Bar(
                name="Kağıt Kaybı",
                y=kesim_labels,
                x=waste_display,
                orientation="h",
                marker_color=WASTE_CLR,
                marker_line_color="white",
                marker_line_width=2,
                text=waste_txt,
                textposition="inside",
                textfont=dict(size=9, color="white", family="Inter, sans-serif"),
                textangle=0,
                insidetextanchor="middle",
                customdata=[[wa] for wa in waste_actual],
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "Kağıt Kaybı: %{customdata[0]:g} cm<br>"
                    "Kağıt kaybı: Evet"
                    "<extra></extra>"
                ),
                showlegend=True,
            ))

            # ── Dummy legend entries for each type ─────────────────────────
            for t in unique_types_used:
                _leg_label = _inv_type_map_anal.get(t, str(t))
                fig_cut.add_trace(go.Bar(
                    name=_leg_label,
                    y=[None], x=[None],
                    orientation="h",
                    marker_color=cut_type_colors.get(t, "#6b7280"),
                    showlegend=True,
                    hoverinfo="skip",
                ))

            # ── Layout — spacious and presentation-quality ──────────────────
            bar_px = 38   # taller bars for readability
            gap_px = 16   # breathing room between rows
            # Dynamic top margin: legend wraps horizontally, ~4 items per row,
            # each row ~28px tall. Add 50px base buffer so legend never overlaps bars.
            _n_leg_items = len(unique_types_used) + 1   # types + waste trace
            _leg_rows    = max(1, math.ceil(_n_leg_items / 4))
            _t_margin    = max(90, _leg_rows * 30 + 55)
            fig_cut.update_layout(
                barmode="stack",
                bargap=0.38,
                height=max(500, n_reels * (bar_px + gap_px) + _t_margin + 80),
                plot_bgcolor=CLR_GRID,
                paper_bgcolor="white",
                xaxis_title="Uzunluk (cm)",
                yaxis_title="",
                legend=dict(
                    orientation="h",
                    yanchor="bottom", y=1.02,
                    xanchor="left", x=0,
                    font=dict(size=11),
                    traceorder="normal",
                    tracegroupgap=4,
                ),
                margin=dict(l=20, r=20, t=_t_margin, b=40),
                xaxis=dict(
                    range=[0, L_val * 1.08],
                    gridcolor="white",
                    tickfont=dict(size=10),
                ),
                yaxis=dict(
                    autorange="reversed",
                    gridcolor="white",
                    tickfont=dict(size=10),
                ),
                uniformtext_minsize=8,
                uniformtext_mode="hide",
            )
            fig_cut.add_vline(
                x=L_val, line_dash="dash", line_color=WASTE_CLR, line_width=1.5,
                annotation_text=f"L = {fmt_n(L_val)} cm",
                annotation_font_color=WASTE_CLR,
                annotation_position="top right",
            )
            st.plotly_chart(fig_cut, use_container_width=True)

        except Exception as e:
            st.error(f"Kesim diyagramı oluşturulamadı: {e}")

    # ── Navigation ─────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    col_exp, _ = st.columns([1, 3])
    with col_exp:
        if st.button("💾  Dışa Aktarma Sayfasına Git →", type="primary", use_container_width=True):
            nav_to("export"); st.rerun()

# MAIN ROUTING
# ===========================================================================
def main():
    
    
    # Check default page
    if "active_page" not in st.session_state:
        st.session_state.active_page = "home"
        
    render_sidebar()
        
    page = st.session_state.active_page
    
    # Routing logic
    if page == "home":
        page_home()
    elif page == "optimization":
        page_optimization()
    elif page == "orders":
        # Legacy redirect → optimization step 1
        st.session_state.optimization_step = 1
        st.session_state.active_page = "optimization"
        st.rerun()
    elif page == "settings":
        # Legacy redirect → optimization step 3
        st.session_state.optimization_step = 3
        st.session_state.active_page = "optimization"
        st.rerun()
    elif page == "results":
        page_results()
    elif page == "analysis":
        page_analysis()
    elif page == "export":
        page_export()
    else:
        st.session_state.active_page = "home"
        page_home()

if __name__ == "__main__":
    main()
